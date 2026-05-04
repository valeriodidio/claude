"""
Export Excel del report Actual Turnover.
Crea un file .xlsx multi-sheet formattato con openpyxl:
  - "Riepilogo"        (tabella principale per macrocat + totale)
  - "Marketplace"      (ripartizione per nomemktp)
  - "Categorie"        (drill cate2 aggregato)
  - "Marche"           (top marche aggregato)
  - "Trend giornaliero"

Con `build_xlsx_split` produce invece un singolo foglio pivot
(righe=macrocat, colonne=marketplace) rispettando le colonne
attualmente espanse nella UI e la metrica compatta scelta.
"""
from __future__ import annotations

from io import BytesIO
from typing import List, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.engine import Engine

from app.services import turnover_query as q
from app.services.turnover_query import Filters


# ---- tabelle metadati metriche (usate anche dal pivot) ----------------------

_METRIC_LABEL = {
    "turnover":     "Fatturato",
    "qta":          "Pezzi",
    "prezzo_medio": "Prezzo medio",
    "costo":        "Costo",
    "margine_pct":  "Margine %",
}
_ALL_METRICS = ["qta", "prezzo_medio", "turnover", "costo", "margine_pct"]


# ---- stili riusabili --------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FONT = Font(bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _write_header(ws, headers: List[str], row: int = 1):
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[row].height = 22


def _autosize(ws, min_w: int = 10, max_w: int = 40):
    """Larghezza colonne adattata al contenuto con min/max."""
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        longest = max((len(str(c.value)) if c.value is not None else 0
                       for c in column_cells), default=10)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, longest + 2))


def _write_data_rows(ws, rows, columns, start_row: int = 2,
                     total: dict | None = None) -> int:
    """Scrive le righe e opzionalmente una riga totale in fondo. Ritorna l'ultima riga."""
    r = start_row
    for item in rows:
        for col_idx, (key, _label, fmt) in enumerate(columns, start=1):
            val = item.get(key)
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = _BORDER
            if fmt == "int":
                cell.number_format = "#,##0"
                cell.alignment = _RIGHT
            elif fmt == "money":
                cell.number_format = '#,##0.00" \u20ac"'
                cell.alignment = _RIGHT
            elif fmt == "pct":
                # lascio il valore percentuale grezzo (es. 23.45) con % a mano
                if val is None:
                    cell.value = "-"
                    cell.alignment = _CENTER
                else:
                    cell.number_format = "0.00\"%\""
                    cell.alignment = _RIGHT
            else:
                cell.alignment = _LEFT
        r += 1

    if total:
        for col_idx, (key, _label, fmt) in enumerate(columns, start=1):
            val = total.get(key, "")
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            if col_idx == 1 and (val in (None, "")):
                cell.value = "TOTALE"
            if fmt == "int":
                cell.number_format = "#,##0"
                cell.alignment = _RIGHT
            elif fmt == "money":
                cell.number_format = '#,##0.00" \u20ac"'
                cell.alignment = _RIGHT
            elif fmt == "pct":
                if val is None:
                    cell.value = "-"
                    cell.alignment = _CENTER
                else:
                    cell.number_format = "0.00\"%\""
                    cell.alignment = _RIGHT
            else:
                cell.alignment = _LEFT
        r += 1

    return r


def _add_summary_sheet(wb: Workbook, engine: Engine, f: Filters):
    ws = wb.active
    ws.title = "Riepilogo"

    # info filtri in cima
    ws["A1"] = "Report Actual Turnover"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Periodo: {f.dal.isoformat()}  \u2192  {f.al.isoformat()}"
    filters_desc = []
    if f.marketplace: filters_desc.append(f"Marketplace: {', '.join(f.marketplace)}")
    if f.fornitore:   filters_desc.append(f"Fornitore: {', '.join(f.fornitore)}")
    if f.macrocat:    filters_desc.append(f"Macro cat.: {', '.join(f.macrocat)}")
    if f.cate2:       filters_desc.append(f"Cat.2: {', '.join(f.cate2)}")
    if f.cate3:       filters_desc.append(f"Cat.3: {', '.join(f.cate3)}")
    if f.marca:       filters_desc.append(f"Marca: {', '.join(f.marca)}")
    if f.codice:      filters_desc.append(f"Codice: {f.codice}")
    ws["A3"] = " | ".join(filters_desc) if filters_desc else "Nessun filtro aggiuntivo"
    ws["A3"].font = Font(italic=True, color="666666")

    data = q.get_summary(engine, f)

    columns = [
        ("macrocat",      "Macro Categoria", "text"),
        ("qta",           "Pezzi venduti",   "int"),
        ("prezzo_medio",  "Prezzo medio",    "money"),
        ("turnover",      "Fatturato",       "money"),
        ("costo",         "Costo",           "money"),
        ("margine_pct",   "Margine %",       "pct"),
    ]
    _write_header(ws, [c[1] for c in columns], row=5)
    last = _write_data_rows(ws, data["rows"], columns, start_row=6, total=data["total"])
    _autosize(ws)


def _add_marketplace_sheet(wb: Workbook, engine: Engine, f: Filters):
    ws = wb.create_sheet("Marketplace")
    data = q.get_marketplace_breakdown(engine, f)
    columns = [
        ("nomemktp",      "Marketplace",   "text"),
        ("qta",           "Pezzi venduti", "int"),
        ("prezzo_medio",  "Prezzo medio",  "money"),
        ("turnover",      "Fatturato",     "money"),
        ("costo",         "Costo",         "money"),
        ("margine_pct",   "Margine %",     "pct"),
    ]
    _write_header(ws, [c[1] for c in columns])
    _write_data_rows(ws, data["rows"], columns)
    _autosize(ws)


def _add_drilldown_sheet(wb: Workbook, engine: Engine, f: Filters,
                         level: str, sheet_name: str, header_label: str):
    ws = wb.create_sheet(sheet_name)
    data = q.get_drilldown(engine, level, f)
    columns = [
        (level,           header_label,    "text"),
        ("qta",           "Pezzi venduti", "int"),
        ("prezzo_medio",  "Prezzo medio",  "money"),
        ("turnover",      "Fatturato",     "money"),
        ("costo",         "Costo",         "money"),
        ("margine_pct",   "Margine %",     "pct"),
    ]
    _write_header(ws, [c[1] for c in columns])
    _write_data_rows(ws, data["rows"], columns)
    _autosize(ws)


def _add_trend_sheet(wb: Workbook, engine: Engine, f: Filters):
    ws = wb.create_sheet("Trend giornaliero")
    data = q.get_trend(engine, "day", f)
    columns = [
        ("bucket",        "Data",          "text"),
        ("qta",           "Pezzi venduti", "int"),
        ("prezzo_medio",  "Prezzo medio",  "money"),
        ("turnover",      "Fatturato",     "money"),
        ("costo",         "Costo",         "money"),
        ("margine_pct",   "Margine %",     "pct"),
    ]
    _write_header(ws, [c[1] for c in columns])
    _write_data_rows(ws, data["rows"], columns)
    _autosize(ws)


def build_xlsx(engine: Engine, f: Filters) -> bytes:
    """Costruisce il workbook e restituisce i bytes."""
    wb = Workbook()
    _add_summary_sheet(wb, engine, f)
    _add_marketplace_sheet(wb, engine, f)
    _add_drilldown_sheet(wb, engine, f, "cate2", "Categorie", "Categoria")
    _add_drilldown_sheet(wb, engine, f, "marca", "Marche", "Marca")
    _add_trend_sheet(wb, engine, f)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Export in modalit\u00e0 "Split per marketplace" (pivot)
# ---------------------------------------------------------------------------

def _filters_desc(f: Filters) -> str:
    parts = []
    if f.marketplace: parts.append(f"Marketplace: {', '.join(f.marketplace)}")
    if f.fornitore:   parts.append(f"Fornitore: {', '.join(f.fornitore)}")
    if f.macrocat:    parts.append(f"Macro cat.: {', '.join(f.macrocat)}")
    if f.cate2:       parts.append(f"Cat.2: {', '.join(f.cate2)}")
    if f.cate3:       parts.append(f"Cat.3: {', '.join(f.cate3)}")
    if f.marca:       parts.append(f"Marca: {', '.join(f.marca)}")
    if f.codice:      parts.append(f"Codice: {f.codice}")
    return " | ".join(parts) if parts else "Nessun filtro aggiuntivo"


def _write_pivot_cell(ws, r: int, c: int, value, metric_key: str,
                      is_total: bool = False):
    cell = ws.cell(row=r, column=c, value=value)
    cell.border = _BORDER
    if is_total:
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
    if metric_key == "qta":
        cell.number_format = "#,##0"
        cell.alignment = _RIGHT
    elif metric_key == "margine_pct":
        if value is None:
            cell.value = "-"
            cell.alignment = _CENTER
        else:
            cell.number_format = "0.00\"%\""
            cell.alignment = _RIGHT
    else:
        cell.number_format = '#,##0.00" \u20ac"'
        cell.alignment = _RIGHT


def _write_pivot_header_cell(ws, r: int, c: int, value: str):
    cell = ws.cell(row=r, column=c, value=value)
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER
    cell.border = _BORDER


def build_xlsx_split(engine: Engine, f: Filters,
                     expanded_mktps: Optional[Set[str]] = None,
                     metric: str = "turnover") -> bytes:
    """Workbook mono-sheet che riproduce la tabella pivot attualmente visibile.

    - `expanded_mktps`: insieme dei nomi di marketplace per cui mostrare
      tutte e 5 le metriche. La chiave speciale '__TOTAL__' espande la
      colonna Totale.
    - `metric`: metrica mostrata nelle colonne compatte (default turnover).
    """
    expanded = set(expanded_mktps or ())
    if metric not in _METRIC_LABEL:
        metric = "turnover"

    data = q.get_summary_by_marketplace(engine, f)
    mktps: List[str] = list(data.get("marketplaces") or [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo split"

    # Intestazione descrittiva (righe 1-4)
    ws["A1"] = "Report Actual Turnover \u2014 Split per marketplace"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Periodo: {f.dal.isoformat()} \u2192 {f.al.isoformat()}"
    ws["A3"] = _filters_desc(f)
    ws["A3"].font = Font(italic=True, color="666666")
    ws["A4"] = f"Metrica compatta: {_METRIC_LABEL[metric]}"
    ws["A4"].font = Font(italic=True, color="666666")

    any_expanded = any(m in expanded for m in mktps) or "__TOTAL__" in expanded
    h1 = 6                       # prima riga intestazione
    h2 = h1 + 1 if any_expanded else h1

    # --- costruzione intestazione (merge rowspan/colspan) --------------------
    col = 1
    _write_pivot_header_cell(ws, h1, col, "Macro categoria")
    if any_expanded:
        ws.merge_cells(start_row=h1, start_column=col, end_row=h2, end_column=col)
    col += 1

    # mappa colonna -> chiave metrica (per formattare correttamente i valori)
    col_meta: List[dict] = []   # [{"kind": "mktp"/"total", "mktp": name, "metric": key}]

    for m in mktps:
        if m in expanded:
            ws.merge_cells(start_row=h1, start_column=col,
                           end_row=h1, end_column=col + len(_ALL_METRICS) - 1)
            _write_pivot_header_cell(ws, h1, col, m)
            for idx, mk in enumerate(_ALL_METRICS):
                _write_pivot_header_cell(ws, h2, col + idx, _METRIC_LABEL[mk])
                col_meta.append({"kind": "mktp", "mktp": m, "metric": mk})
            col += len(_ALL_METRICS)
        else:
            _write_pivot_header_cell(ws, h1, col, m)
            if any_expanded:
                ws.merge_cells(start_row=h1, start_column=col,
                               end_row=h2, end_column=col)
            col_meta.append({"kind": "mktp", "mktp": m, "metric": metric})
            col += 1

    if "__TOTAL__" in expanded:
        ws.merge_cells(start_row=h1, start_column=col,
                       end_row=h1, end_column=col + len(_ALL_METRICS) - 1)
        _write_pivot_header_cell(ws, h1, col, "Totale")
        for idx, mk in enumerate(_ALL_METRICS):
            _write_pivot_header_cell(ws, h2, col + idx, _METRIC_LABEL[mk])
            col_meta.append({"kind": "total", "mktp": None, "metric": mk})
        col += len(_ALL_METRICS)
    else:
        _write_pivot_header_cell(ws, h1, col, "Totale")
        if any_expanded:
            ws.merge_cells(start_row=h1, start_column=col,
                           end_row=h2, end_column=col)
        col_meta.append({"kind": "total", "mktp": None, "metric": metric})
        col += 1

    # altezza righe intestazione
    ws.row_dimensions[h1].height = 22
    if any_expanded:
        ws.row_dimensions[h2].height = 18

    # --- righe dati (TOTALE in testa, poi ogni macrocat) --------------------

    def _write_body_row(row_out: int, label: str, row_data: dict,
                        is_total: bool = False):
        # colonna 1 = etichetta
        cell = ws.cell(row=row_out, column=1, value=label)
        cell.border = _BORDER
        cell.alignment = _LEFT
        if is_total:
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
        cells_by_mktp = row_data.get("cells") or {}
        tot = row_data.get("total") or {}
        for idx, meta in enumerate(col_meta, start=2):
            if meta["kind"] == "mktp":
                src = cells_by_mktp.get(meta["mktp"]) or {}
            else:
                src = tot
            _write_pivot_cell(ws, row_out, idx, src.get(meta["metric"]),
                              meta["metric"], is_total=is_total)

    data_start = h2 + 1
    _write_body_row(data_start, "TOTALE",
                    data.get("total_row") or {}, is_total=True)
    r = data_start + 1
    for item in data.get("rows") or []:
        _write_body_row(r, item.get("macrocat") or "", item, is_total=False)
        r += 1

    # larghezze colonne: prima colonna ampia per i nomi macrocat
    ws.column_dimensions["A"].width = 28
    for ci in range(2, col):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # freeze della prima colonna e delle intestazioni
    ws.freeze_panes = ws.cell(row=data_start, column=2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Export del DETTAGLIO PRODOTTI (tab dinamico della UI)
# ---------------------------------------------------------------------------

# cap di sicurezza: oltre questo numero di righe l'export si ferma
_PRODUCT_EXPORT_CAP = 20000
_PRODUCT_EXPORT_PAGE_SIZE = 2000


def _fetch_all_products_flat(engine: Engine, f: Filters,
                             sort_by: str, sort_dir: str) -> list[dict]:
    rows: list[dict] = []
    page = 1
    while True:
        data = q.get_product_list(
            engine, f,
            page=page, page_size=_PRODUCT_EXPORT_PAGE_SIZE,
            sort_by=sort_by, sort_dir=sort_dir,
        )
        batch = data.get("rows") or []
        rows.extend(batch)
        if (len(batch) < _PRODUCT_EXPORT_PAGE_SIZE
                or page >= (data.get("pages") or 0)
                or len(rows) >= _PRODUCT_EXPORT_CAP):
            break
        page += 1
    return rows[:_PRODUCT_EXPORT_CAP]


def _fetch_all_products_split(engine: Engine, f: Filters,
                              sort_by: str, sort_dir: str
                              ) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    mktps_seen: list[str] = []
    mktps_set: set[str] = set()
    page = 1
    while True:
        data = q.get_product_list_by_marketplace(
            engine, f,
            page=page, page_size=_PRODUCT_EXPORT_PAGE_SIZE,
            sort_by=sort_by, sort_dir=sort_dir,
        )
        batch = data.get("rows") or []
        rows.extend(batch)
        for m in data.get("marketplaces") or []:
            if m not in mktps_set:
                mktps_set.add(m)
                mktps_seen.append(m)
        if (len(batch) < _PRODUCT_EXPORT_PAGE_SIZE
                or page >= (data.get("pages") or 0)
                or len(rows) >= _PRODUCT_EXPORT_CAP):
            break
        page += 1
    return rows[:_PRODUCT_EXPORT_CAP], mktps_seen


def _product_header_lines(ctx: dict, f: Filters, extra: Optional[str] = None
                          ) -> list[tuple[str, bool]]:
    """Ritorna (testo, bold) per le righe di intestazione descrittiva."""
    lines: list[tuple[str, bool]] = []
    lines.append(("Report Actual Turnover \u2014 Dettaglio prodotti", True))
    lines.append(
        (f"Periodo: {f.dal.isoformat()} \u2192 {f.al.isoformat()}", False)
    )
    crumbs: list[str] = []
    for k in ("macrocat", "cate2", "cate3", "marca"):
        v = ctx.get(k)
        if v:
            crumbs.append(f"{v}")
    lines.append(
        (f"Contesto: {' > '.join(crumbs) if crumbs else 'Tutti i prodotti'}",
         False)
    )
    lines.append((_filters_desc(f), False))
    if extra:
        lines.append((extra, False))
    return lines


def build_xlsx_product_list(engine: Engine, f: Filters,
                            ctx: Optional[dict] = None,
                            sort_by: str = "turnover",
                            sort_dir: str = "desc") -> bytes:
    """Export flat del dettaglio prodotti: singolo foglio con tutte le colonne UI."""
    ctx = ctx or {}
    rows = _fetch_all_products_flat(engine, f, sort_by=sort_by, sort_dir=sort_dir)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dettaglio prodotti"

    # intestazione descrittiva
    for i, (txt, bold) in enumerate(_product_header_lines(ctx, f), start=1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = Font(bold=True, size=14) if bold else Font(italic=True, color="666666")

    columns = [
        ("codice",          "Codice",         "text"),
        ("nome",             "Prodotto",       "text"),
        ("marca",            "Marca",          "text"),
        ("macrocat",         "Macro",          "text"),
        ("cate2",            "Categoria",      "text"),
        ("cate3",            "Sotto-cat.",     "text"),
        ("fornitori",        "Fornitori",      "text"),
        ("marketplaces",     "Marketplace",    "text"),
        ("qta",              "Pezzi",          "int"),
        ("ordini",           "Ordini",         "int"),
        ("giorni_attivita",  "Giorni venduti", "int"),
        ("prezzo_medio",     "Prezzo medio",   "money"),
        ("turnover",         "Fatturato",      "money"),
        ("costo",            "Costo",          "money"),
        ("margine_pct",      "Margine %",      "pct"),
        ("first_order",      "Primo ordine",   "text"),
        ("last_order",       "Ultimo ordine",  "text"),
    ]

    header_row = 7
    _write_header(ws, [c[1] for c in columns], row=header_row)
    _write_data_rows(ws, rows, columns, start_row=header_row + 1)

    _autosize(ws)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_xlsx_product_list_split(engine: Engine, f: Filters,
                                  ctx: Optional[dict] = None,
                                  expanded_mktps: Optional[Set[str]] = None,
                                  metric: str = "turnover",
                                  sort_by: str = "turnover",
                                  sort_dir: str = "desc") -> bytes:
    """Export pivot del dettaglio prodotti: struttura identica alla UI."""
    ctx = ctx or {}
    expanded = set(expanded_mktps or ())
    if metric not in _METRIC_LABEL:
        metric = "turnover"

    rows, mktps = _fetch_all_products_split(
        engine, f, sort_by=sort_by, sort_dir=sort_dir,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Dettaglio prodotti split"

    # intestazione descrittiva
    extra = f"Metrica compatta: {_METRIC_LABEL[metric]}"
    for i, (txt, bold) in enumerate(
            _product_header_lines(ctx, f, extra=extra), start=1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = Font(bold=True, size=14) if bold else Font(italic=True, color="666666")

    any_expanded = any(m in expanded for m in mktps) or "__TOTAL__" in expanded
    h1 = 7
    h2 = h1 + 1 if any_expanded else h1

    # prima colonna: Codice / Prodotto
    col = 1
    _write_pivot_header_cell(ws, h1, col, "Codice")
    if any_expanded:
        ws.merge_cells(start_row=h1, start_column=col, end_row=h2, end_column=col)
    col += 1
    _write_pivot_header_cell(ws, h1, col, "Prodotto")
    if any_expanded:
        ws.merge_cells(start_row=h1, start_column=col, end_row=h2, end_column=col)
    col += 1

    # meta colonne: (kind, mktp, metric_key)
    col_meta: list[dict] = []
    for m in mktps:
        if m in expanded:
            ws.merge_cells(start_row=h1, start_column=col,
                           end_row=h1, end_column=col + len(_ALL_METRICS) - 1)
            _write_pivot_header_cell(ws, h1, col, m)
            for idx, mk in enumerate(_ALL_METRICS):
                _write_pivot_header_cell(ws, h2, col + idx, _METRIC_LABEL[mk])
                col_meta.append({"kind": "mktp", "mktp": m, "metric": mk})
            col += len(_ALL_METRICS)
        else:
            _write_pivot_header_cell(ws, h1, col, m)
            if any_expanded:
                ws.merge_cells(start_row=h1, start_column=col,
                               end_row=h2, end_column=col)
            col_meta.append({"kind": "mktp", "mktp": m, "metric": metric})
            col += 1

    if "__TOTAL__" in expanded:
        ws.merge_cells(start_row=h1, start_column=col,
                       end_row=h1, end_column=col + len(_ALL_METRICS) - 1)
        _write_pivot_header_cell(ws, h1, col, "Totale")
        for idx, mk in enumerate(_ALL_METRICS):
            _write_pivot_header_cell(ws, h2, col + idx, _METRIC_LABEL[mk])
            col_meta.append({"kind": "total", "mktp": None, "metric": mk})
        col += len(_ALL_METRICS)
    else:
        _write_pivot_header_cell(ws, h1, col, "Totale")
        if any_expanded:
            ws.merge_cells(start_row=h1, start_column=col,
                           end_row=h2, end_column=col)
        col_meta.append({"kind": "total", "mktp": None, "metric": metric})
        col += 1

    ws.row_dimensions[h1].height = 22
    if any_expanded:
        ws.row_dimensions[h2].height = 18

    # righe dati: una per prodotto. Il "totale" della riga = valore su r stesso.
    data_start = h2 + 1
    for r_idx, row in enumerate(rows, start=data_start):
        # codice
        c = ws.cell(row=r_idx, column=1, value=row.get("codice") or "")
        c.border = _BORDER; c.alignment = _LEFT
        c.font = Font(name="Consolas", size=10)
        # nome
        c = ws.cell(row=r_idx, column=2, value=row.get("nome") or "")
        c.border = _BORDER; c.alignment = _LEFT

        cells_by_mktp = row.get("cells") or {}
        for off, meta in enumerate(col_meta, start=3):
            if meta["kind"] == "mktp":
                src = cells_by_mktp.get(meta["mktp"]) or {}
                val = src.get(meta["metric"])
            else:
                # totale riga = metrica aggregata direttamente su `row`
                val = row.get(meta["metric"])
            _write_pivot_cell(ws, r_idx, off, val, meta["metric"], is_total=False)

    # larghezze
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    for ci in range(3, col):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = ws.cell(row=data_start, column=3)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
