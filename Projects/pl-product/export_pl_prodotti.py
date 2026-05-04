#!/usr/bin/env python3
"""
export_pl_prodotti.py
Esporta il P&L di tutti i prodotti con almeno una vendita negli ultimi 180 giorni.
Output: Excel colorato con le stesse metriche della pagina prodotto-pl-page.asp

Dipendenze:
    pip install pandas sqlalchemy pymysql openpyxl
"""

import sys
import datetime
import os
import time

try:
    import pandas as pd
    from sqlalchemy import create_engine, text
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"\nERRORE import: {e}")
    print("Installa le dipendenze con:\n  pip install pandas sqlalchemy pymysql openpyxl\n")
    sys.exit(1)

# â”€â”€â”€ Configurazione â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
DB_URL        = "mysql+pymysql://valerio:v4l3r10!!!@db-master.yeppon.it:3306/smart2"

# Snapshot multi-periodo: ogni esecuzione salva su DB i risultati per
# ognuno di questi periodi. Lo stesso giorno (data_snapshot=CURDATE())
# convive con la chiave (data_snapshot, id_p, periodo_giorni).
PERIODI       = [30, 60, 90, 120, 180, 360]

# Periodo "principale" usato per generare il file Excel (Dashboard inclusa).
# Deve essere uno dei valori in PERIODI.
PERIODO_EXCEL = 180

# Variabile globale aggiornata nel loop per ciascun periodo.
# Mantenuta per retrocompatibilitÃ  con il footer dell'Excel e i template SQL.
PERIODO       = PERIODO_EXCEL

OUTPUT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    f"pl_prodotti_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# â”€â”€â”€ Query SQL (compatibile MySQL 5.7+, no CTE) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Logica identica all'ASP prodotto-pl-page.asp:
#   - tot_fatturato  = SUM(venduto_tot + spese_sped_grat)
#   - tot_margine    = SUM(utile - comm_mktp - comm_pagamento)
#   - spedizioni     = calcolate per ordine (MAX per ordine poi sommate per prod)
#                      esclude Digitec/Digitec DE
#   - num_resi       = calcolato in Python come SUM(quantita_rma) dal dataframe resi
#   - margine_effettivo = tot_margine + delta_spedizioni
#     (l'impatto economico dei resi non Ã¨ incluso: richiederebbe query molto piÃ¹ lente)

SQL_TPL = """
SELECT
  p.id_p,
  p.codice,
  p.nome,
  IFNULL(m.nome, '')              AS marca,
  IFNULL(p.disp_fornitore, 0)     AS disp_fornitore,
  p.status                         AS status_prodotto,
  IF(pb.id IS NOT NULL AND (pb.dataFine IS NULL OR pb.dataFine >= CURRENT_DATE), 1, 0) AS bloccato,
  pb.dataFine                      AS bloccato_fino,
  metrics.num_ordini,
  metrics.tot_pezzi,
  metrics.tot_fatturato,
  COALESCE(margine_b.tot_margine, metrics.tot_margine) AS tot_margine,
  IFNULL(metrics.categoria,  '')  AS categoria,
  IFNULL(metrics.categoria2, '')  AS categoria2,
  IFNULL(metrics.categoria3, '')  AS categoria3,
  IFNULL(metrics.sender,     '')  AS sender,
  IFNULL(metrics.fornitore,  '')  AS fornitore

FROM (
  -- Metriche principali aggregate per prodotto
  SELECT
    b.id_p,
    COUNT(DISTINCT b.id_ordine)                                                 AS num_ordini,
    IFNULL(SUM(b.quantita), 0)                                                  AS tot_pezzi,
    IFNULL(SUM(b.venduto_tot + b.spese_sped_grat), 0)                           AS tot_fatturato,
    IFNULL(SUM(b.utile - b.comm_mktp - b.comm_pagamento), 0)                    AS tot_margine,
    MAX(b.categoria)                                                             AS categoria,
    MAX(b.categoria2)                                                            AS categoria2,
    MAX(b.categoria3)                                                            AS categoria3,
    MAX(b.sender)                                                                AS sender,
    MAX(b.fornitore)                                                             AS fornitore
  FROM yeppon_stats.bollettato_total AS b
  JOIN ordini_cliente AS oc ON oc.id = b.id_ordine AND oc.pixmania NOT IN (10, 108)
  WHERE b.prezzo_vendita > 0
    AND b.data_evasione >= CURRENT_DATE - INTERVAL {periodo} DAY
  GROUP BY b.id_p
  HAVING tot_fatturato > 0

) AS metrics

LEFT JOIN (
  -- Margine da smart2.bollettato con utile_admin (fallback utile)
  SELECT
    b2.id_p,
    IFNULL(SUM(COALESCE(b2.utile_admin, b2.utile) - b2.comm_mktp - b2.comm_pagamento), 0) AS tot_margine
  FROM smart2.bollettato AS b2
  JOIN ordini_cliente AS oc2 ON oc2.id = b2.id_ordine AND oc2.pixmania NOT IN (10, 108)
  WHERE b2.prezzo_vendita > 0
    AND b2.data_evasione >= CURRENT_DATE - INTERVAL {periodo} DAY
  GROUP BY b2.id_p
) AS margine_b ON margine_b.id_p = metrics.id_p

JOIN prodotti AS p  ON p.id_p  = metrics.id_p AND p.id_c1 <= 20
LEFT JOIN marca   AS m  ON m.id_m  = p.id_m
LEFT JOIN prodotti_bloccati AS pb
       ON pb.codice = p.codice
      AND (pb.dataFine IS NULL OR pb.dataFine >= CURRENT_DATE)
      AND pb.disable_flag = 0

ORDER BY metrics.tot_fatturato DESC
"""

# â”€â”€â”€ Query Resi: legge da tabella pre-calcolata resi_impatto_economico â”€â”€â”€â”€â”€â”€â”€â”€
# La tabella viene popolata da export_dettaglio_resi_completo.py con logica
# completa (AEH, allocation_prodotti, NDC, recovery per tipo, claims, ecc.)
# Filtra per data_rma per essere coerente con il PERIODO di analisi.
SQL_RESI_PRECOMP_TPL = """
SELECT
  id_p,
  id_ordine,
  tipo_rma,
  marketplace AS mktp,
  quantita_rma,
  perdita_netta
FROM yeppon_stats.resi_impatto_economico
WHERE data_rma >= CURRENT_DATE - INTERVAL {periodo} DAY
"""

# â”€â”€â”€ Tabella di destinazione per i dati PL prodotto â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# CREATE TABLE da eseguire una sola volta sul DB:
#
# CREATE TABLE IF NOT EXISTS yeppon_stats.pl_prodotti (
#   id            BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
#   id_p          INT UNSIGNED NOT NULL,
#   codice        VARCHAR(50)  NOT NULL DEFAULT '',
#   nome          VARCHAR(255) NOT NULL DEFAULT '',
#   marca         VARCHAR(100) NOT NULL DEFAULT '',
#   disp_fornitore INT NOT NULL DEFAULT 0,
#   status_prodotto TINYINT NOT NULL DEFAULT 0,
#   bloccato      TINYINT NOT NULL DEFAULT 0,
#   bloccato_fino DATE DEFAULT NULL,
#   num_ordini    INT UNSIGNED NOT NULL DEFAULT 0,
#   tot_pezzi     INT UNSIGNED NOT NULL DEFAULT 0,
#   tot_fatturato DOUBLE NOT NULL DEFAULT 0,
#   tot_margine   DOUBLE NOT NULL DEFAULT 0,
#   perc_margine  DOUBLE NOT NULL DEFAULT 0,
#   tot_sped_fatt DOUBLE NOT NULL DEFAULT 0,
#   tot_sped_costi DOUBLE NOT NULL DEFAULT 0,
#   delta_sped    DOUBLE NOT NULL DEFAULT 0,
#   perc_delta_sped DOUBLE NOT NULL DEFAULT 0,
#   qty_resi      INT UNSIGNED NOT NULL DEFAULT 0,
#   perc_resi     DOUBLE NOT NULL DEFAULT 0,
#   imp_eco_resi  DOUBLE NOT NULL DEFAULT 0,
#   imp_Danni     DOUBLE NOT NULL DEFAULT 0,
#   imp_Danno_rientrato DOUBLE NOT NULL DEFAULT 0,
#   imp_Difettosi DOUBLE NOT NULL DEFAULT 0,
#   imp_Giacenza  DOUBLE NOT NULL DEFAULT 0,
#   imp_Mancato_Ritiro DOUBLE NOT NULL DEFAULT 0,
#   imp_Prodotto_non_conforme DOUBLE NOT NULL DEFAULT 0,
#   imp_Recesso   DOUBLE NOT NULL DEFAULT 0,
#   imp_Reclamo_contestazioni DOUBLE NOT NULL DEFAULT 0,
#   imp_Smarrimenti DOUBLE NOT NULL DEFAULT 0,
#   margine_effettivo DOUBLE NOT NULL DEFAULT 0,
#   perc_margine_eff  DOUBLE NOT NULL DEFAULT 0,
#   periodo_giorni INT UNSIGNED NOT NULL DEFAULT 180,
#   data_calcolo  DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
#   INDEX idx_id_p (id_p),
#   INDEX idx_codice (codice),
#   INDEX idx_data_calcolo (data_calcolo)
# ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;

PL_TABLE = "yeppon_stats.pl_prodotti"
PL_DB_COLS = [
    "id_p", "codice", "nome", "marca", "disp_fornitore",
    "categoria", "categoria2", "categoria3", "sender", "fornitore",
    "status_prodotto", "bloccato", "bloccato_fino",
    "num_ordini", "tot_pezzi", "tot_fatturato", "tot_margine", "perc_margine",
    "tot_sped_fatt", "tot_sped_costi", "delta_sped", "perc_delta_sped",
    "qty_resi", "perc_resi", "imp_eco_resi",
    "imp_Danni", "imp_Danno_rientrato", "imp_Difettosi", "imp_Giacenza",
    "imp_Mancato_Ritiro", "imp_Prodotto_non_conforme", "imp_Recesso",
    "imp_Reclamo_contestazioni", "imp_Smarrimenti",
    "margine_effettivo", "perc_margine_eff",
    "periodo_giorni",
]

# Colonna snapshot per separare i salvataggi nel tempo (gg/mm/aaaa).
# Insieme a periodo_giorni e id_p forma la chiave logica della riga.
PL_SNAPSHOT_COL = "data_snapshot"

# Tabella di destinazione per il dettaglio spedizioni per (prodotto, corriere).
# Vengono inserite SOLO le righe con fattura corriere reale; gli ordini privi
# di fattura sono ignorati (no fallback tariffario qui).
PL_CORR_TABLE = "yeppon_stats.pl_prodotti_corrieri"
PL_CORR_DB_COLS = [
    "id_p", "corriere", "num_ordini", "tot_pezzi",
    "tot_sped_fatt", "tot_sped_costi", "delta_sped", "perc_delta_sped",
    "periodo_giorni",
]

# Tabella di dettaglio per ordine: snapshot per (data_snapshot, periodo_giorni).
# Strategia DELETE+INSERT come pl_prodotti / pl_prodotti_corrieri.
PL_CORR_DET_TABLE = "yeppon_stats.pl_prodotti_corrieri_dettaglio"
PL_CORR_DET_DB_COLS = [
    "id_p", "id_ordine", "corriere",
    "quantita", "sped_fatt", "sped_costi", "delta_sped",
    "periodo_giorni",
]

# Tabella per breakdown vendite/margine per (id_p, marketplace).
# Snapshot per (data_snapshot, periodo_giorni): DELETE+INSERT idempotente.
PL_MKTP_TABLE = "yeppon_stats.pl_prodotti_marketplace"
PL_MKTP_DB_COLS = [
    "id_p", "marketplace",
    "num_ordini", "tot_pezzi", "tot_fatturato",
    "tot_margine", "perc_margine",
    "tot_sped_fatt", "tot_sped_costi", "delta_sped", "perc_delta_sped",
    "qty_resi", "perc_resi", "imp_eco_resi",
    "margine_effettivo", "perc_margine_eff",
    "periodo_giorni",
]


# â”€â”€â”€ Query per ricalcolo spedizioni corretto â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Dati grezzi per-ordine/per-prodotto con pesi da product_dimension e prodotti.
# In Italia si usa GREATEST(peso, weight_volume), all'estero peso effettivo.

# â”€â”€â”€ Query per metriche base per marketplace â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
SQL_PER_MKTP_TPL = """
SELECT
  metrics.id_p,
  metrics.marketplace,
  metrics.num_ordini,
  metrics.tot_pezzi,
  metrics.tot_fatturato,
  COALESCE(margine_b.tot_margine, metrics.tot_margine) AS tot_margine
FROM (
  SELECT
    b.id_p,
    b.mktp                                                     AS marketplace,
    COUNT(DISTINCT b.id_ordine)                                 AS num_ordini,
    IFNULL(SUM(b.quantita), 0)                                  AS tot_pezzi,
    IFNULL(SUM(b.venduto_tot + b.spese_sped_grat), 0)           AS tot_fatturato,
    IFNULL(SUM(b.utile - b.comm_mktp - b.comm_pagamento), 0)    AS tot_margine
  FROM yeppon_stats.bollettato_total AS b
  JOIN ordini_cliente AS oc ON oc.id = b.id_ordine AND oc.pixmania NOT IN (10, 108)
  WHERE b.prezzo_vendita > 0
    AND b.data_evasione >= CURRENT_DATE - INTERVAL {periodo} DAY
  GROUP BY b.id_p, b.mktp
  HAVING tot_fatturato > 0
) AS metrics
LEFT JOIN (
  SELECT
    b2.id_p,
    b2.mktp AS marketplace,
    IFNULL(SUM(COALESCE(b2.utile_admin, b2.utile) - b2.comm_mktp - b2.comm_pagamento), 0) AS tot_margine
  FROM smart2.bollettato AS b2
  JOIN ordini_cliente AS oc2 ON oc2.id = b2.id_ordine AND oc2.pixmania NOT IN (10, 108)
  WHERE b2.prezzo_vendita > 0
    AND b2.data_evasione >= CURRENT_DATE - INTERVAL {periodo} DAY
  GROUP BY b2.id_p, b2.mktp
) AS margine_b ON margine_b.id_p = metrics.id_p AND margine_b.marketplace = metrics.marketplace
"""

SQL_SHIPPING_PRICES = "SELECT shipping_type_id, weight_from, weight_to, price AS cost FROM shipping_price"

SQL_SPESE_GRATUITE = "SELECT spese_gratuite FROM impo_generali LIMIT 1"

SQL_COUNTRY_STYPE = """
SELECT LOWER(country) AS country, MIN(shipping_type_id) AS shipping_type_id
FROM shipping_price
GROUP BY LOWER(country)
"""

SQL_SPED_RAW_TPL = """
SELECT
  b2.id_ordine,
  b2.id_p,
  SUM(b2.quantita)                                          AS quantita,
  MAX(b2.mktp)                                               AS mktp,
  MAX(b2.spese_sped_grat + b2.spese_sped_cliente)           AS sped_rev_ordine,
  GREATEST(IFNULL(pd.peso, IFNULL(p.peso, 0)), 0)           AS peso_prodotto,
  IFNULL(pd.weight_volume, 0)                                AS peso_volume,
  MAX(oc2.id_spe)                                            AS id_spe,
  MAX(oc2.importo_spe)                                       AS importo_spe,
  MAX(oc2.pixmania)                                          AS pixmania,
  LOWER(LEFT(IFNULL(MAX(b2.stato), 'IT'), 2))               AS country,
  SUM(b2.venduto_tot)                                        AS prezzo_vendita,
  IFNULL(MAX(b2.spese_sped_corriere), 0)                     AS spese_sped_corriere
FROM yeppon_stats.bollettato_total AS b2
JOIN ordini_cliente AS oc2
     ON oc2.id = b2.id_ordine AND oc2.pixmania NOT IN (10, 108)
LEFT JOIN product_dimension AS pd ON pd.id_p = b2.id_p
LEFT JOIN prodotti          AS p  ON p.id_p  = b2.id_p
WHERE b2.prezzo_vendita > 0
  AND b2.data_evasione >= CURRENT_DATE - INTERVAL {periodo} DAY
GROUP BY b2.id_ordine, b2.id_p
"""

# â”€â”€â”€ Query costi spedizione reali dai corrieri â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Il costo spedizione viene letto direttamente dalle fatture dei corrieri
# (imponibile) nelle tabelle fornitori.*, poi riproporzionato per prodotto
# in base al peso.
# Ogni corriere viene interrogato separatamente per evitare blocchi.

_COURIER_QUERIES = [
    ("TNT", """
        SELECT id_ordine, imponibile
        FROM fornitori.tnt_spedizioni
        WHERE id_ordine IN (SELECT id_ordine FROM tmp_pl_orders)
    """),
    ("DHL", """
        SELECT dhls.id_ordine,
               ROUND(dhls.imponibile + IFNULL(dhlt.imponibile, 0), 2) AS imponibile
        FROM fornitori.dhl_spedizioni AS dhls
        LEFT JOIN (
            SELECT YEAR(d2.data_fattura) AS anno, MONTH(d2.data_fattura) AS mese,
                (d2.imponibile_trazione + d2.importo_resi + d2.importo_extramisura
                 + d2.importo_correzione_indirizzo + d2.imponibile_express)
                / COUNT(d1.ldv) AS imponibile
            FROM fornitori.dhl_spedizioni d1
            JOIN fornitori.dhl_trazioni d2
              ON YEAR(d2.data_fattura) = YEAR(d1.data_fattura)
             AND MONTH(d2.data_fattura) = MONTH(d1.data_fattura)
            GROUP BY YEAR(d2.data_fattura), MONTH(d2.data_fattura)
        ) AS dhlt
          ON dhlt.anno = YEAR(dhls.data_fattura) AND dhlt.mese = MONTH(dhls.data_fattura)
        WHERE dhls.id_ordine IN (SELECT id_ordine FROM tmp_pl_orders)
    """),
    ("Fastest", """
        SELECT riferimento AS id_ordine, SUM(imponibile) AS imponibile
        FROM fornitori.fastest_spedizioni
        WHERE riferimento IN (SELECT id_ordine FROM tmp_pl_orders)
        GROUP BY riferimento
    """),
    ("BRT", """
        SELECT id_ordine, imponibile
        FROM fornitori.dpd_spedizioni
        WHERE id_ordine IN (SELECT id_ordine FROM tmp_pl_orders)
    """),
    ("Zust", """
        SELECT order_id AS id_ordine, SUM(imponibile) AS imponibile
        FROM fornitori.zust_spedizioni
        WHERE order_id IN (SELECT id_ordine FROM tmp_pl_orders)
        GROUP BY order_id
    """),
    ("Arcese", """
        SELECT RifMittente AS id_ordine, SUM(Imponibile) AS imponibile
        FROM fornitori.arcese_spedizioni
        WHERE RifMittente IN (SELECT id_ordine FROM tmp_pl_orders)
        GROUP BY RifMittente
    """),
    ("Milkman", """
        SELECT id_ordine, SUM(imponibile) AS imponibile
        FROM fornitori.milkman_spedizioni
        WHERE id_ordine IN (SELECT id_ordine FROM tmp_pl_orders)
        GROUP BY id_ordine
    """),
    ("Fercam", """
        SELECT id_ordine, SUM(imponibile) AS imponibile
        FROM fornitori.fercam_spedizioni
        WHERE id_ordine IN (SELECT id_ordine FROM tmp_pl_orders)
        GROUP BY id_ordine
    """),
    ("GLS", """
        SELECT id_ordine, SUM(imponibile) AS imponibile
        FROM fornitori.gls_spedizioni
        WHERE id_ordine IN (SELECT id_ordine FROM tmp_pl_orders)
        GROUP BY id_ordine
    """),
    ("Spring", """
        SELECT id_ordine, SUM(imponibile) AS imponibile
        FROM fornitori.spring_spedizioni
        WHERE id_ordine IN (SELECT id_ordine FROM tmp_pl_orders)
        GROUP BY id_ordine
    """),
]

# SDA viene gestita separatamente: scarica tutto e fa match in Python
# (il LIKE '%id' senza indice Ã¨ troppo lento in SQL su grandi volumi)
SQL_SDA_ALL = """
    SELECT rif_spedizione, imponibile
    FROM fornitori.sda_spedizioni
    WHERE imponibile IS NOT NULL AND imponibile > 0
"""


# â”€â”€â”€ Funzioni di calcolo spedizioni (replica logica ASP) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def assicurasped_py(peso: float, prezzo: float, costosped: float) -> float:
    """Replica ASP assicurasped: aggiunge costo assicurazione."""
    if peso > 10 or prezzo > 1000:
        if peso < 400:
            return costosped + 1
        else:
            return costosped + (prezzo * 0.6 / 100)
    return costosped


def trova_real_costo_spedizione_py(
    shipping_type_id: int,
    peso: float,
    prezzo: float,
    spese_gratuite: float,
    sp_lookup: dict,
) -> float:
    """
    Replica ASP trova_real_costo_spedizione.
    Restituisce il costo spedizione (IVA inclusa).
    sp_lookup: dict {shipping_type_id: [(weight_from, weight_to, cost), ...]}
    """
    id_spedizione = shipping_type_id
    id_spe = shipping_type_id

    if id_spedizione == 38:
        id_spe = 32  # stessi costi di TNT

    costo = 0.0
    if id_spe and peso:
        for wf, wt, c in sp_lookup.get(id_spe, []):
            if wf <= peso <= wt:
                costo = c
                break

    if prezzo and prezzo >= spese_gratuite:
        costo = 0.0

    if id_spedizione == 38:
        costo += 5

    if id_spedizione == 36:
        costo = 2.0

    if id_spedizione not in (36, 39, 99, 50, 60, 61):
        costo = assicurasped_py(peso, prezzo, costo)

    if id_spe in (1, 26, 29, 30, 31, 100, 101, 102, 103, 104, 105):
        return 0.0

    return costo


def calc_sped_corrette(
    df_raw: pd.DataFrame,
    sp_df: pd.DataFrame,
    spese_gratuite: float,
    country_stype: dict,
    df_courier_costs: pd.DataFrame = None,
) -> pd.DataFrame:
    """
    Calcola spedizioni corrette per prodotto con pesi coerenti.
    Italia  â†’ GREATEST(peso, weight_volume) da product_dimension
    Estero  â†’ peso effettivo da prodotti
    Riproporziona costo ordine per peso_prodotto / peso_totale_ordine.
    Restituisce DataFrame: id_p, tot_sped_fatt, tot_sped_costi
    """
    if df_raw.empty:
        return pd.DataFrame(columns=["id_p", "tot_sped_fatt", "tot_sped_costi"])

    # Build shipping_price lookup
    sp_lookup: dict = {}
    for _, r in sp_df.iterrows():
        stid = int(r["shipping_type_id"])
        sp_lookup.setdefault(stid, []).append(
            (float(r["weight_from"]), float(r["weight_to"]), float(r["cost"]))
        )

    df = df_raw.copy()
    for col in ["peso_prodotto", "peso_volume", "quantita", "importo_spe",
                "prezzo_vendita", "sped_rev_ordine", "id_spe", "pixmania",
                "spese_sped_corriere"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    df["mktp"] = df["mktp"].astype(str).fillna("")

    # â”€â”€ Peso corretto per prodotto â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    #   Italia  â†’ GREATEST(peso_prodotto, peso_volume)
    #   Estero  â†’ peso_prodotto (peso reale)
    is_it = df["country"] == "it"
    df["peso_corretto"] = df["peso_prodotto"].copy()
    df.loc[is_it, "peso_corretto"] = df.loc[is_it, ["peso_prodotto", "peso_volume"]].max(axis=1)
    df["peso_corretto"] = df["peso_corretto"].clip(lower=0.001)  # min 1 g

    df["peso_pesato"] = df["peso_corretto"] * df["quantita"]

    # â”€â”€ Aggregazioni a livello ordine â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ord_agg = df.groupby("id_ordine").agg(
        peso_tot_ord=("peso_pesato", "sum"),
        pv_tot_ord=("prezzo_vendita", "sum"),
    )
    ord_agg["peso_tot_ord"] = ord_agg["peso_tot_ord"].clip(lower=0.001)
    df = df.merge(ord_agg, on="id_ordine")

    # â”€â”€ Shipping type per ordine (dalla country in shipping_price) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    df["stype"] = df["country"].map(country_stype).fillna(32).astype(int)
    # Override per marketplace/tipo specifico
    mktp_str = df["mktp"].astype(str)
    df.loc[mktp_str == "3", "stype"] = 200         # eBay
    df.loc[df["id_spe"] == 106, "stype"] = 201

    # â”€â”€ Calcolo spedizione a livello ordine â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Ricalcoliamo sia revenue (spese_sped_cliente) che costi per ordine
    # usando pesi corretti (GREATEST per IT, peso reale per estero).
    order_cols = ["id_ordine", "stype", "peso_tot_ord", "pv_tot_ord",
                  "id_spe", "importo_spe", "pixmania", "mktp"]
    orders = df.drop_duplicates(subset=["id_ordine"])[order_cols].copy()

    # sped_rev_map: revenue spedizione ricalcolata per ordine
    # sped_cost_fallback_map: costo calcolato con tariffe (fallback per ordini senza costo corriere reale)
    sped_rev_map: dict = {}
    sped_cost_fallback_map: dict = {}

    for _, o in orders.iterrows():
        id_ordine   = o["id_ordine"]
        importo_spe = float(o["importo_spe"])
        id_spe_val  = int(o["id_spe"])
        pix         = int(o["pixmania"])
        mktp_val    = str(o["mktp"])
        peso_tot    = float(o["peso_tot_ord"])
        pv_tot      = float(o["pv_tot_ord"])
        stype       = int(o["stype"])

        # ── Stima costo da tariffa (calcolata SEMPRE come fallback) ─────────────
        # Usata quando la fattura corriere non è ancora stata caricata.
        # Evita che delta_sped risulti artificialmente positivo a inizio mese
        # per gli ordini il cui costo reale non è ancora disponibile.
        sped_cost_est = 0.0
        if id_spe_val == 111:
            if   peso_tot < 20: sped_cost_est = 22.37
            elif peso_tot < 30: sped_cost_est = 26.2
            elif peso_tot < 40: sped_cost_est = 28.67
            elif peso_tot < 50: sped_cost_est = 35.73
            elif peso_tot < 60: sped_cost_est = 37.84
            else:               sped_cost_est = peso_tot * 0.41
        elif pix == 41 or mktp_val == "57":
            sped_cost_est = 0.0
        elif id_spe_val not in (31, 26, 28):
            sped_cost_est = trova_real_costo_spedizione_py(
                stype, peso_tot, pv_tot, spese_gratuite, sp_lookup
            ) / 1.22

        if importo_spe > 0:
            # Revenue dal fee effettivo incassato dal cliente; costo stimato da tariffa
            # come fallback per quando la fattura corriere non è ancora arrivata.
            sped_rev_map[id_ordine] = importo_spe / 1.22
        else:
            # Revenue e costo entrambi dalla stima tariffaria
            sped_rev_map[id_ordine] = sped_cost_est

        # Il fallback tariffario è SEMPRE disponibile: se la fattura corriere
        # non è stata caricata, usiamo la stima invece di 0.
        sped_cost_fallback_map[id_ordine] = sped_cost_est

    df["sped_rev_ordine_calc"] = df["id_ordine"].map(sped_rev_map).fillna(0)
    df["sped_cost_fallback"] = df["id_ordine"].map(sped_cost_fallback_map)

    # â”€â”€ Proporzione per prodotto (in base al peso) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    df["quota_peso"] = df["peso_pesato"] / df["peso_tot_ord"]

    # Revenue (Incasso Sped.): SEMPRE ricalcolata con pesi corretti
    # Digitec/Digitec DE (Asendia): incasso e costo spedizione = 0
    is_digitec = df["mktp"].astype(str).isin(["digitec", "digitec DE"])
    df["sped_fatt_prod"] = (
        (~is_digitec).astype(float)
        * df["sped_rev_ordine_calc"]
        * df["quota_peso"]
    )

    # Costi (Costo Sped.):
    #   - Digitec/Digitec DE (Asendia) → 0
    #   - Fattura corriere reale disponibile → costo_sped_reale * quota_peso
    #   - Nessuna fattura → fallback tariffario (sped_cost_fallback, ora sempre != None)
    if df_courier_costs is not None and not df_courier_costs.empty:
        df = df.merge(df_courier_costs[["id_ordine", "costo_sped_reale"]],
                      on="id_ordine", how="left")
        df["costo_sped_reale"] = df["costo_sped_reale"].fillna(0)
    else:
        df["costo_sped_reale"] = 0.0

    df["sped_costi_prod"] = 0.0
    not_digitec = ~is_digitec
    has_courier = not_digitec & (df["costo_sped_reale"] > 0)
    # Ordini con fattura corriere reale
    df.loc[has_courier, "sped_costi_prod"] = (
        df.loc[has_courier, "costo_sped_reale"] * df.loc[has_courier, "quota_peso"]
    )
    # Ordini senza fattura corriere: usa la stima tariffaria (sempre disponibile).
    # Garantisce che delta_sped sia significativo anche a inizio mese,
    # quando alcune fatture corrieri non sono ancora state caricate.
    no_courier = not_digitec & (df["costo_sped_reale"] <= 0)
    fallback_available = no_courier & df["sped_cost_fallback"].notna()
    df.loc[fallback_available, "sped_costi_prod"] = (
        df.loc[fallback_available, "sped_cost_fallback"]
        * df.loc[fallback_available, "quota_peso"]
    )
    # fallback_missing è ora sempre vuoto (sped_cost_fallback è sempre impostato),
    # ma manteniamo per robustezza: in caso estremo, costo = 0.
    fallback_missing = no_courier & df["sped_cost_fallback"].isna()
    df.loc[fallback_missing, "sped_costi_prod"] = 0.0

    # â”€â”€ Aggregazione per prodotto â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    result = df.groupby("id_p").agg(
        tot_sped_fatt=("sped_fatt_prod", "sum"),
        tot_sped_costi=("sped_costi_prod", "sum"),
    ).reset_index()

    # Per-ordine: shipping revenue ricalcolata (serve per resi)
    per_ord = df.drop_duplicates(subset=["id_ordine"])[["id_ordine", "sped_rev_ordine_calc"]].copy()

    # Per-(id_p, id_ordine): breakdown necessario per snapshot corrieri
    per_ord_prod = df.groupby(["id_p", "id_ordine"]).agg(
        quantita=("quantita", "sum"),
        sped_fatt_prod=("sped_fatt_prod", "sum"),
        sped_costi_prod=("sped_costi_prod", "sum"),
    ).reset_index()

    # Per-(id_p, mktp): spedizioni divise per marketplace
    per_mktp = df.groupby(["id_p", "mktp"]).agg(
        tot_sped_fatt=("sped_fatt_prod", "sum"),
        tot_sped_costi=("sped_costi_prod", "sum"),
    ).reset_index().rename(columns={"mktp": "marketplace"})

    return {"per_prod": result, "per_ord": per_ord, "per_mktp": per_mktp,
            "per_ord_prod": per_ord_prod}


# â”€â”€â”€ Palette colori Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
F_HEADER = PatternFill("solid", fgColor="1F3864")   # blu scuro header
F_GREEN  = PatternFill("solid", fgColor="C6EFCE")
F_YELLOW = PatternFill("solid", fgColor="FFEB9C")
F_RED    = PatternFill("solid", fgColor="FFC7CE")
F_GREY   = PatternFill("solid", fgColor="F2F2F2")   # righe pari
F_WHITE  = PatternFill("solid", fgColor="FFFFFF")   # righe dispari

T_HEADER = Font(bold=True, color="FFFFFF", size=11)
T_GREEN  = Font(color="276221", size=11)
T_YELLOW = Font(color="9C5700", size=11)
T_RED    = Font(color="9C0006", size=11)
T_NORMAL = Font(color="333333", size=11)
T_NOTE   = Font(italic=True, color="888888", size=10)

A_CENTER = Alignment(horizontal="center", vertical="center")
A_RIGHT  = Alignment(horizontal="right",  vertical="center")
A_LEFT   = Alignment(horizontal="left",   vertical="center")

# â”€â”€â”€ Struttura colonne dell'Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# (field, header, width, fmt, color_type)
# color_type: None | 'margine' | 'margine_eff' | 'delta' | 'resi' | 'imp_resi'
COLS = [
    ("id_p",             "ID Prodotto",           10, "txt",  None),
    ("codice",           "Codice",                14, "str",  None),
    ("nome",             "Nome Prodotto",          50, "str",  None),
    ("marca",            "Marca",                  22, "str",  None),
    ("disp_fornitore",   "Disp. Fornitore",        16, "int",  None),
    ("status_prodotto",  "Status",                 10, "status", None),
    ("bloccato",         "Bloccato",               10, "yn",   None),
    ("bloccato_fino",    "Bloccato Fino",           14, "date", None),
    ("num_ordini",       "NÂ° Ordini",              11, "int",  None),
    ("tot_pezzi",        "Pezzi Venduti",           13, "int",  None),
    ("tot_fatturato",    "Fatturato (â‚¬)",           15, "eur",  None),
    ("tot_margine",      "Margine (â‚¬)",             15, "eur",  "margine"),
    ("perc_margine",     "% Margine",               12, "perc", "margine"),
    ("tot_sped_fatt",    "Incasso Sped. (â‚¬)",        16, "eur",  None),
    ("tot_sped_costi",   "Costo Sped. (â‚¬)",          15, "eur",  None),
    ("delta_sped",       "Delta Sped. (â‚¬)",          15, "eur",  "delta"),
    ("perc_delta_sped",  "% Delta Sped.",            13, "perc", "delta"),
    ("qty_resi",         "Qty Resi",                  10, "int",  None),
    ("perc_resi",        "% Resi",                    9, "perc", "resi"),
    ("imp_eco_resi",              "Impatto Eco. Resi (â‚¬)",          22, "eur",  "imp_resi"),
    ("imp_Danni",                  "Danni (â‚¬)",                      15, "eur",  None),
    ("imp_Danno_rientrato",        "Danno rientrato (â‚¬)",            18, "eur",  None),
    ("imp_Difettosi",              "Difettosi (â‚¬)",                  15, "eur",  None),
    ("imp_Giacenza",               "Giacenza (â‚¬)",                   15, "eur",  None),
    ("imp_Mancato_Ritiro",         "Mancato Ritiro (â‚¬)",             17, "eur",  None),
    ("imp_Prodotto_non_conforme",   "Prodotto non conforme (â‚¬)",     22, "eur",  None),
    ("imp_Recesso",                "Recesso (â‚¬)",                    15, "eur",  None),
    ("imp_Reclamo_contestazioni",   "Reclamo e contestazioni (â‚¬)",   24, "eur",  None),
    ("imp_Smarrimenti",            "Smarrimenti (â‚¬)",                16, "eur",  None),
    ("margine_effettivo",          "Margine Effettivo (â‚¬)",          20, "eur",  "margine_eff"),
    ("perc_margine_eff", "% Margine Effettivo",       20, "perc", "margine_eff"),
]


def cell_color(color_type: str, value: float):
    """Restituisce (fill, font) in base al tipo di metrica e al valore."""
    if color_type in ("margine", "margine_eff"):
        # value = percentuale margine
        if value < 0:    return F_RED,    T_RED
        if value <= 10:  return F_YELLOW, T_YELLOW
        return F_GREEN, T_GREEN
    if color_type == "delta":
        # value = delta spedizioni in â‚¬
        return (F_GREEN, T_GREEN) if value >= 0 else (F_RED, T_RED)
    if color_type == "resi":
        # value = percentuale resi
        if value <= 2:   return F_GREEN,  T_GREEN
        if value <= 5:   return F_YELLOW, T_YELLOW
        return F_RED, T_RED
    if color_type == "imp_resi":
        return (F_RED, T_RED) if value > 0 else (F_GREEN, T_GREEN)
    return None, None


TIPI_RMA = ["Danni", "Danno rientrato", "Difettosi", "Giacenza",
            "Mancato Ritiro", "Prodotto non conforme", "Recesso",
            "Reclamo e contestazioni", "Smarrimenti"]
IMP_RMA_COLS = [f"imp_{t.replace(' ', '_').replace('e_contestazioni', 'contestazioni')}" for t in TIPI_RMA]
# mapping tipo_rma â†’ colonna
TIPO_TO_COL = dict(zip(TIPI_RMA, IMP_RMA_COLS))


def compute(df: pd.DataFrame) -> pd.DataFrame:
    """Calcola le metriche derivate (percentuali, delta, margine effettivo)."""
    base_cols = ["tot_fatturato", "tot_margine", "tot_pezzi",
                 "tot_sped_fatt", "tot_sped_costi", "qty_resi"] + IMP_RMA_COLS
    for col in base_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    nan = float("nan")
    df["delta_sped"]        = df["tot_sped_fatt"]    - df["tot_sped_costi"]
    df["perc_margine"]      = (df["tot_margine"]      / df["tot_fatturato"].replace(0, nan) * 100).round(2).fillna(0)
    df["perc_delta_sped"]   = (df["delta_sped"]        / df["tot_sped_fatt"].replace(0, nan) * 100).round(2).fillna(0)
    df["perc_resi"]         = (df["qty_resi"]           / df["tot_pezzi"].replace(0, nan)    * 100).round(2).fillna(0)
    df["imp_eco_resi"]      = df[IMP_RMA_COLS].sum(axis=1)
    df["margine_effettivo"] = df["tot_margine"] + df["delta_sped"] - df["imp_eco_resi"]
    df["perc_margine_eff"]  = (df["margine_effettivo"] / df["tot_fatturato"].replace(0, nan) * 100).round(2).fillna(0)

    # Campi testuali per nuove colonne (safe se gia' convertiti)
    def _status(x):
        try:
            return "Attivo" if int(x) == 1 else "Disattivo"
        except (ValueError, TypeError):
            return str(x) if x else ""
    def _bloccato(x):
        try:
            return "SÃ¬" if int(x) == 1 else "No"
        except (ValueError, TypeError):
            return str(x) if x else ""
    df["status_prodotto"] = df["status_prodotto"].apply(_status)
    df["bloccato"] = df["bloccato"].apply(_bloccato)
    return df


def _write_sheet(ws, df: pd.DataFrame, sheet_label: str = ""):
    """Formatta un singolo foglio Excel con le colonne COLS."""
    # â”€â”€ Header â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws.row_dimensions[1].height = 44
    for cell in ws[1]:
        cell.fill      = F_HEADER
        cell.font      = T_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # â”€â”€ Larghezze colonne â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for i, (_, _, w, _, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes    = "E2"
    ws.auto_filter.ref = ws.dimensions

    # â”€â”€ Dati riga per riga â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    nrows = len(df)
    for ri in range(2, nrows + 2):
        row_data   = df.iloc[ri - 2]
        base_fill  = F_GREY if ri % 2 == 0 else F_WHITE

        pct_map = {
            "margine":     row_data["perc_margine"],
            "margine_eff": row_data["perc_margine_eff"],
            "delta":       row_data["delta_sped"],
            "resi":        row_data["perc_resi"],
            "imp_resi":    row_data["imp_eco_resi"],
        }

        for ci, (field, _, _, fmt, ctype) in enumerate(COLS, 1):
            cell = ws.cell(row=ri, column=ci)

            if fmt == "eur":
                cell.number_format = 'â‚¬#,##0.00'
                cell.alignment     = A_RIGHT
            elif fmt == "perc":
                cell.number_format = '0.00"%"'
                cell.alignment     = A_RIGHT
            elif fmt == "int":
                cell.number_format = "#,##0"
                cell.alignment     = A_CENTER
            elif fmt == "txt":
                cell.number_format = '@'
                cell.value         = str(cell.value) if cell.value is not None else ""
                cell.alignment     = A_LEFT
            elif fmt in ("status", "yn"):
                cell.alignment = A_CENTER
            elif fmt == "date":
                if cell.value and str(cell.value) not in ('', 'NaT', 'None', 'nan'):
                    cell.number_format = 'DD/MM/YYYY'
                else:
                    cell.value = ""
                cell.alignment = A_CENTER
            else:
                cell.alignment = A_LEFT

            if ctype:
                fill, font = cell_color(ctype, pct_map[ctype])
                cell.fill  = fill
                cell.font  = font
            else:
                cell.fill = base_fill
                cell.font = T_NORMAL

    # â”€â”€ Nota a piÃ¨ foglio â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    note_row = nrows + 3
    label_extra = f" â€” {sheet_label}" if sheet_label else ""
    note = ws.cell(
        row=note_row, column=1,
        value=(
            f"* Dati: ultimi {PERIODO} giorni{label_extra} | "
            "Margine Effettivo = Margine + Delta Spedizioni - Impatto Economico Resi. "
            "Logica identica a prodotto-pl-page.asp."
        )
    )
    note.font = T_NOTE
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row,   end_column=len(COLS)
    )


def _write_dashboard(ws, dash: pd.DataFrame):
    """Formatta il foglio dashboard corrieri."""
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    DASH_COLS = [
        ("corriere",             "Corriere",           20, "str"),
        ("num_ordini",           "NÂ° Ordini",          14, "int"),
        ("incasso_sped",         "Incasso Sped. (â‚¬)",  18, "eur"),
        ("costo_sped",           "Costo Sped. (â‚¬)",    18, "eur"),
        ("delta",                "Delta (â‚¬)",          16, "eur"),
        ("perc_delta",           "% Delta",            12, "perc"),
        ("perc_costo_sul_totale","% sul Totale Costi", 18, "perc"),
    ]

    # Header
    ws.row_dimensions[1].height = 36
    for ci, (_, hdr, w, _) in enumerate(DASH_COLS, 1):
        cell = ws.cell(row=1, column=ci)
        cell.value = hdr
        cell.fill = F_HEADER
        cell.font = T_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    eur_fmt  = '#,##0.00 â‚¬'
    perc_fmt = '0.00"%"'
    int_fmt  = '#,##0'
    tot_idx  = len(dash)  # 1-based last data row = tot_idx + 1

    for ri, row in dash.iterrows():
        excel_row = ri + 2  # 1-based, skip header
        is_total = str(row.get("corriere", "")) == "TOTALE"
        for ci, (field, _, _, fmt) in enumerate(DASH_COLS, 1):
            cell = ws.cell(row=excel_row, column=ci)
            val = row[field]
            cell.value = val
            if fmt == "eur":
                cell.number_format = eur_fmt
                cell.alignment = A_RIGHT
            elif fmt == "perc":
                cell.number_format = perc_fmt
                cell.alignment = A_RIGHT
            elif fmt == "int":
                cell.number_format = int_fmt
                cell.alignment = A_RIGHT
            else:
                cell.alignment = A_LEFT

            if is_total:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill("solid", fgColor="D9E2F3")
            elif field == "delta":
                if val >= 0:
                    cell.fill = F_GREEN
                    cell.font = Font(color="006100")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006")
            elif field == "perc_delta":
                if val >= 0:
                    cell.font = Font(color="006100")
                else:
                    cell.font = Font(color="9C0006")

    # â”€â”€ Grafici â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    data_rows = len(dash) - 1  # escludi TOTALE
    if data_rows >= 1:
        # Grafico a barre: incasso vs costo per corriere
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Incasso vs Costo Spedizione per Corriere"
        bar.y_axis.title = "Euro (â‚¬)"
        bar.x_axis.title = "Corriere"
        bar.width = 28
        bar.height = 16
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + data_rows)
        incasso_data = Reference(ws, min_col=3, min_row=1, max_row=1 + data_rows)
        costo_data   = Reference(ws, min_col=4, min_row=1, max_row=1 + data_rows)
        bar.add_data(incasso_data, titles_from_data=True)
        bar.add_data(costo_data, titles_from_data=True)
        bar.set_categories(cats)
        bar.shape = 4
        ws.add_chart(bar, "I2")

        # Grafico a torta: ripartizione costi
        pie = PieChart()
        pie.title = "Ripartizione Costi per Corriere"
        pie.width = 20
        pie.height = 16
        pie_cats = Reference(ws, min_col=1, min_row=2, max_row=1 + data_rows)
        pie_data = Reference(ws, min_col=4, min_row=1, max_row=1 + data_rows)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_cats)
        ws.add_chart(pie, "I20")


def to_excel(df: pd.DataFrame, path: str, mktp_dfs: dict = None,
             dashboard_corrieri: pd.DataFrame = None):
    """Genera il file Excel con foglio master + un foglio per marketplace + dashboard."""
    fields  = [c[0] for c in COLS]
    headers = [c[1] for c in COLS]
    out = df[fields].copy()
    out.columns = headers

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # â”€â”€ Foglio master â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        out.to_excel(writer, sheet_name="P&L Prodotti", index=False)
        wb = writer.book
        ws = writer.sheets["P&L Prodotti"]
        print(f"   Formattazione master ({len(df)} righe)...", end="", flush=True)
        _write_sheet(ws, df)
        print(" OK")

        # â”€â”€ Fogli per marketplace â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if mktp_dfs:
            for mktp_name, mktp_df in mktp_dfs.items():
                if mktp_df.empty:
                    continue
                # Nome foglio: max 31 char, no caratteri speciali
                sheet_name = str(mktp_name)[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "(").replace("]", ")")
                mktp_out = mktp_df[fields].copy()
                mktp_out.columns = headers
                mktp_out.to_excel(writer, sheet_name=sheet_name, index=False)
                ws_mktp = writer.sheets[sheet_name]
                print(f"   Formattazione {sheet_name} ({len(mktp_df)} righe)...", end="", flush=True)
                _write_sheet(ws_mktp, mktp_df, sheet_label=mktp_name)
                print(" OK")

        # â”€â”€ Dashboard corrieri â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if dashboard_corrieri is not None and not dashboard_corrieri.empty:
            dash_fields = ["corriere", "num_ordini", "incasso_sped", "costo_sped",
                           "delta", "perc_delta", "perc_costo_sul_totale"]
            dashboard_corrieri[dash_fields].to_excel(
                writer, sheet_name="Dashboard Corrieri", index=False)
            ws_dash = writer.sheets["Dashboard Corrieri"]
            print(f"   Formattazione Dashboard Corrieri...", end="", flush=True)
            _write_dashboard(ws_dash, dashboard_corrieri)
            print(" OK")


def _run_period(periodo: int, engine, sp_df, spese_gratuite, country_stype, t0):
    """
    Esegue l'intera pipeline di calcolo per un singolo periodo (giorni)
    e salva il risultato come snapshot su yeppon_stats.pl_prodotti.

    Restituisce (df, mktp_dfs, dash) per uso esterno (es. generazione Excel).
    """
    global PERIODO
    PERIODO = periodo  # per nota footer Excel e log

    print(f"\n--- PERIODO {periodo} giorni ---")
    print("Query metriche principali...", end="", flush=True)
    df = pd.read_sql(text(SQL_TPL.format(periodo=periodo)), engine)
    print(f" {len(df)} prodotti ({time.time() - t0:.1f}s)")

    print("Query dati spedizioni per ordine/prodotto...", end="", flush=True)
    df_sped_raw = pd.read_sql(text(SQL_SPED_RAW_TPL.format(periodo=periodo)), engine)
    print(f" {len(df_sped_raw)} righe ({time.time() - t0:.1f}s)")

    print("Query costi corrieri reali...", flush=True)
    order_ids = df_sped_raw["id_ordine"].unique().tolist()
    order_ids_set = set(int(x) for x in order_ids)
    courier_frames = []
    with engine.connect() as conn:
        conn.execute(text("DROP TABLE IF EXISTS tmp_pl_orders"))
        conn.execute(text(
            "CREATE TABLE tmp_pl_orders ("
            "id_ordine BIGINT PRIMARY KEY"
            ") ENGINE=MEMORY"
        ))
        batch_size = 5000
        for i in range(0, len(order_ids), batch_size):
            batch = order_ids[i:i + batch_size]
            vals = ",".join(f"({int(x)})" for x in batch)
            conn.execute(text(f"INSERT INTO tmp_pl_orders (id_ordine) VALUES {vals}"))
        conn.commit()

        for courier_name, courier_sql in _COURIER_QUERIES:
            print(f"   {courier_name}...", end="", flush=True)
            try:
                cdf = pd.read_sql(text(courier_sql), conn)
                cdf = cdf[cdf["imponibile"].notna() & (cdf["imponibile"] > 0)]
                if not cdf.empty:
                    cdf = cdf[["id_ordine", "imponibile"]].copy()
                    cdf["corriere"] = courier_name
                    courier_frames.append(cdf)
                print(f" {len(cdf)} righe")
            except Exception as e:
                print(f" ERRORE: {e}")

        # SDA: scarica tutto e fa match set-based in Python
        print("   SDA...", end="", flush=True)
        try:
            sda_df = pd.read_sql(text(SQL_SDA_ALL), conn)
            if not sda_df.empty:
                from collections import defaultdict
                oid_by_len = defaultdict(set)
                for oid in order_ids_set:
                    oid_by_len[len(str(oid))].add(str(oid))
                sda_df["rif_spedizione"] = sda_df["rif_spedizione"].astype(str)
                sda_matches = []
                for ln, oid_set in oid_by_len.items():
                    suffixes = sda_df["rif_spedizione"].str[-ln:]
                    mask = suffixes.isin(oid_set)
                    if mask.any():
                        matched = sda_df.loc[mask].copy()
                        matched["id_ordine"] = suffixes[mask].astype(int)
                        sda_matches.append(matched[["id_ordine", "imponibile"]])
                if sda_matches:
                    sda_result = pd.concat(sda_matches, ignore_index=True)
                    sda_result["corriere"] = "SDA"
                    courier_frames.append(sda_result)
                    print(f" {len(sda_result)} righe")
                else:
                    print(" 0 righe")
            else:
                print(" 0 righe")
        except Exception as e:
            print(f" ERRORE: {e}")

        conn.execute(text("DROP TABLE IF EXISTS tmp_pl_orders"))
        conn.commit()

    if courier_frames:
        df_courier_all = pd.concat(courier_frames, ignore_index=True)
        df_courier_all["id_ordine"] = pd.to_numeric(df_courier_all["id_ordine"], errors="coerce")
        df_courier_all["imponibile"] = pd.to_numeric(df_courier_all["imponibile"], errors="coerce")
        n_before = len(df_courier_all)
        df_courier_all = df_courier_all.drop_duplicates(subset=["id_ordine", "corriere", "imponibile"])
        n_dropped = n_before - len(df_courier_all)
        if n_dropped:
            print(f"   Dedup: rimosse {n_dropped} righe duplicate (stesso ordine/corriere/importo)")
        df_courier_costs = df_courier_all.groupby("id_ordine")["imponibile"].sum().reset_index()
        df_courier_costs.columns = ["id_ordine", "costo_sped_reale"]
        df_courier_names = df_courier_all.sort_values("imponibile", ascending=False)\
            .drop_duplicates(subset=["id_ordine"], keep="first")[["id_ordine", "corriere"]]
    else:
        df_courier_costs = pd.DataFrame(columns=["id_ordine", "costo_sped_reale"])
        df_courier_names = pd.DataFrame(columns=["id_ordine", "corriere"])
    print(f"   Totale: {len(df_courier_costs)} ordini con costi ({time.time() - t0:.1f}s)")

    print("Ricalcolo spedizioni (pesi corretti)...", end="", flush=True)
    sped_data = calc_sped_corrette(df_sped_raw, sp_df, spese_gratuite, country_stype, df_courier_costs)
    df_sped = sped_data["per_prod"]
    sped_per_mktp = sped_data["per_mktp"]
    print(f" {len(df_sped)} prodotti ({time.time() - t0:.1f}s)")

    # â”€â”€ Resi (filtrati per data_rma coerente con il PERIODO) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    print("Query dati resi (da resi_impatto_economico)...", end="", flush=True)
    try:
        df_resi_raw = pd.read_sql(
            text(SQL_RESI_PRECOMP_TPL.format(periodo=periodo)), engine
        )
        print(f" {len(df_resi_raw)} righe ({time.time() - t0:.1f}s)")
    except Exception as e:
        print(f"\n  ATTENZIONE: query resi fallita ({e}).")
        df_resi_raw = pd.DataFrame(columns=["id_p", "id_ordine", "tipo_rma",
                                            "mktp", "quantita_rma", "perdita_netta"])

    print("Aggregazione impatto resi...", end="", flush=True)
    if not df_resi_raw.empty:
        for col in ["quantita_rma", "perdita_netta"]:
            df_resi_raw[col] = pd.to_numeric(df_resi_raw[col], errors="coerce").fillna(0)
        df_resi_raw["mktp"] = df_resi_raw["mktp"].astype(str).fillna("")
        df_resi_raw["tipo_rma"] = df_resi_raw["tipo_rma"].astype(str).fillna("")

        df_resi_raw["imp_col"] = df_resi_raw["tipo_rma"].map(TIPO_TO_COL)
        imp_pivot = df_resi_raw.dropna(subset=["imp_col"]).pivot_table(
            index="id_p", columns="imp_col", values="perdita_netta",
            aggfunc="sum", fill_value=0).reset_index()
        for c in IMP_RMA_COLS:
            if c not in imp_pivot.columns:
                imp_pivot[c] = 0.0

        qty_resi = df_resi_raw.groupby("id_p")["quantita_rma"].sum().reset_index()
        qty_resi.columns = ["id_p", "qty_resi"]
    else:
        imp_pivot = pd.DataFrame(columns=["id_p"] + IMP_RMA_COLS)
        qty_resi = pd.DataFrame(columns=["id_p", "qty_resi"])

    df_resi = imp_pivot.merge(qty_resi, on="id_p", how="outer").fillna(0)
    print(f" {len(df_resi)} prodotti con resi ({time.time() - t0:.1f}s)")

    # â”€â”€ Merge risultati â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    df = df.merge(df_resi, on="id_p", how="left")
    for _c in IMP_RMA_COLS:
        df[_c] = df[_c].fillna(0)
    df["qty_resi"] = df["qty_resi"].fillna(0).astype(int)

    df = df.merge(df_sped, on="id_p", how="left")
    df["tot_sped_fatt"] = df["tot_sped_fatt"].fillna(0)
    df["tot_sped_costi"] = df["tot_sped_costi"].fillna(0)

    df = compute(df)

    # â”€â”€ Dati per marketplace â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    print("Query metriche per marketplace...", end="", flush=True)
    df_mktp = pd.read_sql(text(SQL_PER_MKTP_TPL.format(periodo=periodo)), engine)
    marketplaces = sorted(df_mktp["marketplace"].dropna().unique())
    print(f" {len(df_mktp)} righe, {len(marketplaces)} marketplace ({time.time() - t0:.1f}s)")

    prod_info_cols = ["id_p", "codice", "nome", "marca", "disp_fornitore",
                      "categoria", "categoria2", "categoria3", "sender", "fornitore",
                      "status_prodotto", "bloccato", "bloccato_fino"]
    df_prod_info = df[prod_info_cols].copy()

    print("Aggregazione resi per marketplace...", end="", flush=True)
    if not df_resi_raw.empty:
        imp_pivot_m = df_resi_raw.dropna(subset=["imp_col"]).pivot_table(
            index=["id_p", "mktp"], columns="imp_col", values="perdita_netta",
            aggfunc="sum", fill_value=0).reset_index().rename(columns={"mktp": "marketplace"})
        for c in IMP_RMA_COLS:
            if c not in imp_pivot_m.columns:
                imp_pivot_m[c] = 0.0
        qty_resi_m = df_resi_raw.groupby(["id_p", "mktp"])["quantita_rma"].sum().reset_index()
        qty_resi_m.columns = ["id_p", "marketplace", "qty_resi"]
        resi_per_mktp = imp_pivot_m.merge(qty_resi_m, on=["id_p", "marketplace"], how="outer").fillna(0)
    else:
        resi_per_mktp = pd.DataFrame(columns=["id_p", "marketplace"] + IMP_RMA_COLS + ["qty_resi"])
    print(f" OK ({time.time() - t0:.1f}s)")

    mktp_dfs = {}
    for mktp_name in marketplaces:
        m = df_mktp[df_mktp["marketplace"] == mktp_name][["id_p", "num_ordini", "tot_pezzi", "tot_fatturato", "tot_margine"]].copy()
        m = m.merge(df_prod_info, on="id_p", how="left")
        m = m.merge(sped_per_mktp[sped_per_mktp["marketplace"] == mktp_name].drop(columns=["marketplace"]),
                    on="id_p", how="left")
        m = m.merge(resi_per_mktp[resi_per_mktp["marketplace"] == mktp_name].drop(columns=["marketplace"]),
                    on="id_p", how="left")
        m["tot_sped_fatt"] = m["tot_sped_fatt"].fillna(0)
        m["tot_sped_costi"] = m["tot_sped_costi"].fillna(0)
        for _c in IMP_RMA_COLS:
            m[_c] = m[_c].fillna(0)
        m["qty_resi"] = m["qty_resi"].fillna(0).astype(int)
        m = compute(m)
        m = m.sort_values("tot_fatturato", ascending=False)
        mktp_dfs[str(mktp_name)] = m

    # Salvataggio snapshot marketplace su DB (per app) con metriche complete
    _save_snapshot_marketplace(mktp_dfs, periodo, engine, t0)

    # â”€â”€ Dashboard corrieri â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    print("Costruzione dashboard corrieri...", end="", flush=True)
    df_sped_ord = df_sped_raw.drop_duplicates(subset=["id_ordine"])[["id_ordine", "mktp"]].copy()
    df_sped_ord["mktp"] = df_sped_ord["mktp"].astype(str).fillna("")
    is_dig_ord = df_sped_ord["mktp"].isin(["digitec", "digitec DE"])
    df_sped_ord = df_sped_ord[~is_dig_ord]
    df_sped_ord = df_sped_ord.merge(df_courier_names, on="id_ordine", how="left")
    df_sped_ord["corriere"] = df_sped_ord["corriere"].fillna("Tariffa (fallback)")
    df_sped_ord = df_sped_ord.merge(df_courier_costs, on="id_ordine", how="left")
    df_sped_ord["costo_sped_reale"] = df_sped_ord["costo_sped_reale"].fillna(0)
    df_sped_ord = df_sped_ord.merge(sped_data["per_ord"], on="id_ordine", how="left")
    df_sped_ord["sped_rev_ordine_calc"] = df_sped_ord["sped_rev_ordine_calc"].fillna(0)
    is_fallback = df_sped_ord["corriere"] == "Tariffa (fallback)"
    df_sped_ord.loc[is_fallback, "costo_sped_reale"] = df_sped_ord.loc[is_fallback, "sped_rev_ordine_calc"]
    dash = df_sped_ord.groupby("corriere").agg(
        num_ordini=("id_ordine", "nunique"),
        incasso_sped=("sped_rev_ordine_calc", "sum"),
        costo_sped=("costo_sped_reale", "sum"),
    ).reset_index()
    dash["delta"] = dash["incasso_sped"] - dash["costo_sped"]
    dash["perc_delta"] = (dash["delta"] / dash["incasso_sped"].replace(0, float("nan")) * 100).round(2).fillna(0)
    dash["perc_costo_sul_totale"] = (dash["costo_sped"] / dash["costo_sped"].sum() * 100).round(2)
    dash = dash.sort_values("costo_sped", ascending=False).reset_index(drop=True)
    tot_row = pd.DataFrame([{
        "corriere": "TOTALE",
        "num_ordini": dash["num_ordini"].sum(),
        "incasso_sped": dash["incasso_sped"].sum(),
        "costo_sped": dash["costo_sped"].sum(),
        "delta": dash["delta"].sum(),
        "perc_delta": (dash["delta"].sum() / dash["incasso_sped"].sum() * 100) if dash["incasso_sped"].sum() != 0 else 0,
        "perc_costo_sul_totale": 100.0,
    }])
    dash = pd.concat([dash, tot_row], ignore_index=True)
    print(f" {len(dash) - 1} corrieri ({time.time() - t0:.1f}s)")

    # â”€â”€ Snapshot pl_prodotti_corrieri â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Aggrega (id_p, corriere) usando SOLO ordini con fattura reale
    # (inner join con df_courier_names). Gli ordini senza fattura caricata
    # sono volutamente esclusi.
    print("Aggregazione (id_p, corriere)...", end="", flush=True)
    per_ord_prod = sped_data["per_ord_prod"]
    df_corr = per_ord_prod.merge(df_courier_names, on="id_ordine", how="inner")
    if df_corr.empty:
        df_corrieri = pd.DataFrame(columns=PL_CORR_DB_COLS)
    else:
        df_corrieri = df_corr.groupby(["id_p", "corriere"]).agg(
            num_ordini=("id_ordine", "nunique"),
            tot_pezzi=("quantita", "sum"),
            tot_sped_fatt=("sped_fatt_prod", "sum"),
            tot_sped_costi=("sped_costi_prod", "sum"),
        ).reset_index()
        df_corrieri["delta_sped"] = (
            df_corrieri["tot_sped_fatt"] - df_corrieri["tot_sped_costi"]
        ).round(2)
        df_corrieri["perc_delta_sped"] = (
            df_corrieri["delta_sped"]
            / df_corrieri["tot_sped_fatt"].replace(0, pd.NA)
            * 100
        ).round(2).fillna(0)
        for c in ["tot_sped_fatt", "tot_sped_costi"]:
            df_corrieri[c] = df_corrieri[c].round(2)
        df_corrieri["tot_pezzi"] = df_corrieri["tot_pezzi"].astype(int)
    print(f" {len(df_corrieri)} righe ({time.time() - t0:.1f}s)")

    # â”€â”€ Salvataggio snapshot su DB â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _save_snapshot(df, periodo, engine, t0)
    _save_snapshot_corrieri(df_corrieri, periodo, engine, t0)
    _save_snapshot_corrieri_dettaglio(df_corr, periodo, engine, t0)

    return df, mktp_dfs, dash


def _save_snapshot(df: pd.DataFrame, periodo: int, engine, t0):
    """
    Salva il dataframe come snapshot su yeppon_stats.pl_prodotti.

    Strategia: per (data_snapshot=CURDATE(), periodo_giorni=periodo) cancella
    le righe esistenti e re-inserisce. Idempotente per re-run nello stesso giorno.
    Le righe di altri data_snapshot/periodo_giorni restano intatte â†’ storico.
    """
    print(f"Salvataggio snapshot DB (periodo={periodo})...", end="", flush=True)

    # Assicura presenza di periodo_giorni nel df sorgente prima della select.
    df = df.copy()
    df["periodo_giorni"] = periodo

    df_db = df[PL_DB_COLS].copy()
    df_db["status_prodotto"] = pd.to_numeric(df["status_prodotto"], errors="coerce").fillna(0).astype(int)
    df_db["bloccato"] = df["bloccato"].apply(lambda x: 1 if str(x).strip().lower() in ("1", "sÃ¬", "si", "yes") else 0)
    df_db["bloccato_fino"] = pd.to_datetime(df["bloccato_fino"], errors="coerce")
    df_db["periodo_giorni"] = periodo

    # Dedup difensiva: la chiave UNIQUE è (data_snapshot, id_p, periodo_giorni).
    # Eventuali duplicati per id_p (es. dovuti a prodotti_bloccati con più
    # blocchi attivi sullo stesso codice) farebbero fallire l'INSERT.
    n_pre = len(df_db)
    df_db = df_db.drop_duplicates(subset=["id_p"], keep="first")
    n_dup = n_pre - len(df_db)
    if n_dup:
        print(f" [dedup {n_dup} duplicati id_p]", end="", flush=True)

    today = datetime.date.today()
    now = datetime.datetime.now()

    # Colonne dell'INSERT: PL_DB_COLS + data_snapshot + data_calcolo
    insert_cols = PL_DB_COLS + [PL_SNAPSHOT_COL, "data_calcolo"]
    placeholders = ", ".join(["%s"] * len(insert_cols))
    col_names = ", ".join([f"`{c}`" for c in insert_cols])
    insert_sql = f"INSERT INTO {PL_TABLE} ({col_names}) VALUES ({placeholders})"

    rows = []
    for _, r in df_db.iterrows():
        vals = []
        for c in PL_DB_COLS:
            v = r[c]
            if pd.isna(v):
                vals.append(None)
            elif isinstance(v, pd.Timestamp):
                vals.append(v.to_pydatetime())
            else:
                vals.append(v)
        vals.append(today)  # data_snapshot
        vals.append(now)    # data_calcolo
        rows.append(tuple(vals))

    with engine.connect() as conn:
        # Cancella eventuali righe giÃ  presenti per (oggi, periodo)
        conn.execute(
            text(
                f"DELETE FROM {PL_TABLE} "
                f"WHERE {PL_SNAPSHOT_COL} = :d AND periodo_giorni = :p"
            ),
            {"d": today, "p": periodo},
        )
        conn.commit()

        raw_conn = conn.connection
        cursor = raw_conn.cursor()
        BATCH = 1000
        for i in range(0, len(rows), BATCH):
            cursor.executemany(insert_sql, rows[i:i + BATCH])
        raw_conn.commit()
        cursor.close()

    print(f" {len(rows)} righe inserite ({time.time() - t0:.1f}s)")


def _save_snapshot_corrieri(df_corr: pd.DataFrame, periodo: int, engine, t0):
    """
    Snapshot dettaglio (id_p, corriere) su yeppon_stats.pl_prodotti_corrieri.
    DELETE+INSERT per (data_snapshot=CURDATE(), periodo_giorni=periodo).
    Idempotente per re-run nello stesso giorno.
    Solo ordini con fattura corriere reale (no fallback tariffario).
    """
    print(f"Salvataggio snapshot corrieri DB (periodo={periodo})...", end="", flush=True)

    if df_corr.empty:
        print(" 0 righe (nessun ordine con fattura corriere)")
        return

    df_corr = df_corr.copy()
    df_corr["periodo_giorni"] = periodo
    df_db = df_corr[PL_CORR_DB_COLS].copy()
    df_db["periodo_giorni"] = periodo

    # Dedup difensiva sulla chiave logica (id_p, corriere)
    n_pre = len(df_db)
    df_db = df_db.drop_duplicates(subset=["id_p", "corriere"], keep="first")
    n_dup = n_pre - len(df_db)
    if n_dup:
        print(f" [dedup {n_dup}]", end="", flush=True)

    today = datetime.date.today()
    now = datetime.datetime.now()

    insert_cols = PL_CORR_DB_COLS + [PL_SNAPSHOT_COL, "data_calcolo"]
    placeholders = ", ".join(["%s"] * len(insert_cols))
    col_names = ", ".join([f"`{c}`" for c in insert_cols])
    insert_sql = f"INSERT INTO {PL_CORR_TABLE} ({col_names}) VALUES ({placeholders})"

    rows = []
    for _, r in df_db.iterrows():
        vals = []
        for c in PL_CORR_DB_COLS:
            v = r[c]
            vals.append(None if pd.isna(v) else v)
        vals.append(today)
        vals.append(now)
        rows.append(tuple(vals))

    with engine.connect() as conn:
        conn.execute(
            text(
                f"DELETE FROM {PL_CORR_TABLE} "
                f"WHERE {PL_SNAPSHOT_COL} = :d AND periodo_giorni = :p"
            ),
            {"d": today, "p": periodo},
        )
        conn.commit()
        raw_conn = conn.connection
        cursor = raw_conn.cursor()
        BATCH = 1000
        for i in range(0, len(rows), BATCH):
            cursor.executemany(insert_sql, rows[i:i + BATCH])
        raw_conn.commit()
        cursor.close()

    print(f" {len(rows)} righe inserite ({time.time() - t0:.1f}s)")


def _save_snapshot_corrieri_dettaglio(df_corr_det: pd.DataFrame, periodo: int, engine, t0):
    """
    Dettaglio per (id_p, id_ordine, corriere) su pl_prodotti_corrieri_dettaglio.
    Snapshot per (data_snapshot=CURDATE(), periodo_giorni=periodo):
    DELETE selettivo + INSERT, idempotente per re-run nello stesso giorno.
    Solo ordini con fattura corriere reale.
    """
    print(f"Salvataggio dettaglio corrieri DB (periodo={periodo})...", end="", flush=True)

    today = datetime.date.today()
    now = datetime.datetime.now()

    if df_corr_det is None or df_corr_det.empty:
        with engine.connect() as conn:
            conn.execute(
                text(
                    f"DELETE FROM {PL_CORR_DET_TABLE} "
                    f"WHERE {PL_SNAPSHOT_COL} = :d AND periodo_giorni = :p"
                ),
                {"d": today, "p": periodo},
            )
            conn.commit()
        print(" 0 righe")
        return

    df_db = df_corr_det.rename(columns={
        "sped_fatt_prod":  "sped_fatt",
        "sped_costi_prod": "sped_costi",
    }).copy()
    df_db["periodo_giorni"] = periodo
    df_db["delta_sped"] = (df_db["sped_fatt"] - df_db["sped_costi"]).round(2)
    df_db["sped_fatt"] = df_db["sped_fatt"].round(2)
    df_db["sped_costi"] = df_db["sped_costi"].round(2)
    df_db["quantita"] = df_db["quantita"].astype(int)

    df_db = df_db[PL_CORR_DET_DB_COLS].drop_duplicates(
        subset=["id_p", "id_ordine"], keep="first"
    )

    insert_cols = PL_CORR_DET_DB_COLS + [PL_SNAPSHOT_COL, "data_calcolo"]
    placeholders = ", ".join(["%s"] * len(insert_cols))
    col_names = ", ".join([f"`{c}`" for c in insert_cols])
    insert_sql = f"INSERT INTO {PL_CORR_DET_TABLE} ({col_names}) VALUES ({placeholders})"

    rows = []
    for _, r in df_db.iterrows():
        vals = []
        for c in PL_CORR_DET_DB_COLS:
            v = r[c]
            vals.append(None if pd.isna(v) else v)
        vals.append(today)
        vals.append(now)
        rows.append(tuple(vals))

    with engine.connect() as conn:
        conn.execute(
            text(
                f"DELETE FROM {PL_CORR_DET_TABLE} "
                f"WHERE {PL_SNAPSHOT_COL} = :d AND periodo_giorni = :p"
            ),
            {"d": today, "p": periodo},
        )
        conn.commit()
        raw_conn = conn.connection
        cursor = raw_conn.cursor()
        BATCH = 5000
        for i in range(0, len(rows), BATCH):
            cursor.executemany(insert_sql, rows[i:i + BATCH])
        raw_conn.commit()
        cursor.close()

    print(f" {len(rows)} righe inserite ({time.time() - t0:.1f}s)")


def _save_snapshot_marketplace(mktp_dfs: dict, periodo: int, engine, t0):
    """
    Snapshot vendite per (id_p, marketplace) su yeppon_stats.pl_prodotti_marketplace.
    DELETE+INSERT per (data_snapshot=CURDATE(), periodo_giorni=periodo).
    Idempotente per re-run nello stesso giorno.

    `mktp_dfs` è il dict {marketplace_name: DataFrame} già passato per
    `compute()`, quindi contiene anche delta spedizioni, resi e margine
    effettivo.
    """
    print(f"Salvataggio snapshot marketplace DB (periodo={periodo})...", end="", flush=True)

    today = datetime.date.today()
    now = datetime.datetime.now()

    if not mktp_dfs:
        with engine.connect() as conn:
            conn.execute(
                text(
                    f"DELETE FROM {PL_MKTP_TABLE} "
                    f"WHERE {PL_SNAPSHOT_COL} = :d AND periodo_giorni = :p"
                ),
                {"d": today, "p": periodo},
            )
            conn.commit()
        print(" 0 righe")
        return

    # Concatena tutti i DataFrame per marketplace, aggiungendo la colonna marketplace
    parts = []
    for mktp_name, m in mktp_dfs.items():
        if m is None or m.empty:
            continue
        tmp = m.copy()
        tmp["marketplace"] = str(mktp_name)
        parts.append(tmp)

    if not parts:
        with engine.connect() as conn:
            conn.execute(
                text(
                    f"DELETE FROM {PL_MKTP_TABLE} "
                    f"WHERE {PL_SNAPSHOT_COL} = :d AND periodo_giorni = :p"
                ),
                {"d": today, "p": periodo},
            )
            conn.commit()
        print(" 0 righe")
        return

    df_db = pd.concat(parts, ignore_index=True)
    df_db["periodo_giorni"] = periodo
    df_db["marketplace"] = df_db["marketplace"].fillna("").astype(str)

    int_cols = ["num_ordini", "tot_pezzi", "qty_resi"]
    for c in int_cols:
        if c in df_db.columns:
            df_db[c] = pd.to_numeric(df_db[c], errors="coerce").fillna(0).astype(int)

    float_cols = [
        "tot_fatturato", "tot_margine", "perc_margine",
        "tot_sped_fatt", "tot_sped_costi", "delta_sped", "perc_delta_sped",
        "perc_resi", "imp_eco_resi", "margine_effettivo", "perc_margine_eff",
    ]
    for c in float_cols:
        if c in df_db.columns:
            df_db[c] = pd.to_numeric(df_db[c], errors="coerce").fillna(0).round(2)
        else:
            df_db[c] = 0.0

    df_db = df_db[PL_MKTP_DB_COLS].drop_duplicates(
        subset=["id_p", "marketplace"], keep="first"
    )

    insert_cols = PL_MKTP_DB_COLS + [PL_SNAPSHOT_COL, "data_calcolo"]
    placeholders = ", ".join(["%s"] * len(insert_cols))
    col_names = ", ".join([f"`{c}`" for c in insert_cols])
    insert_sql = f"INSERT INTO {PL_MKTP_TABLE} ({col_names}) VALUES ({placeholders})"

    rows = []
    for _, r in df_db.iterrows():
        vals = []
        for c in PL_MKTP_DB_COLS:
            v = r[c]
            vals.append(None if pd.isna(v) else v)
        vals.append(today)
        vals.append(now)
        rows.append(tuple(vals))

    with engine.connect() as conn:
        conn.execute(
            text(
                f"DELETE FROM {PL_MKTP_TABLE} "
                f"WHERE {PL_SNAPSHOT_COL} = :d AND periodo_giorni = :p"
            ),
            {"d": today, "p": periodo},
        )
        conn.commit()
        raw_conn = conn.connection
        cursor = raw_conn.cursor()
        BATCH = 1000
        for i in range(0, len(rows), BATCH):
            cursor.executemany(insert_sql, rows[i:i + BATCH])
        raw_conn.commit()
        cursor.close()

    print(f" {len(rows)} righe inserite ({time.time() - t0:.1f}s)")


def main():
    t0 = time.time()
    print()
    print("=" * 60)
    print(f"  P&L Export â€” Snapshot multi-periodo: {PERIODI}")
    print(f"  Excel generato per periodo principale: {PERIODO_EXCEL} giorni")
    print("=" * 60)

    host = DB_URL.split("@")[-1].split("/")[0]
    print(f"\nConnessione a {host}...")
    engine = create_engine(DB_URL, connect_args={"connect_timeout": 60})

    # â”€â”€ Dati comuni a tutti i periodi (caricati una sola volta) â”€â”€â”€â”€â”€â”€â”€â”€â”€
    print("Query tariffe spedizione...", end="", flush=True)
    sp_df = pd.read_sql(text(SQL_SHIPPING_PRICES), engine)
    sg_df = pd.read_sql(text(SQL_SPESE_GRATUITE), engine)
    spese_gratuite = float(sg_df.iloc[0, 0]) if not sg_df.empty else 9999.0
    cs_df = pd.read_sql(text(SQL_COUNTRY_STYPE), engine)
    country_stype = dict(zip(cs_df["country"], cs_df["shipping_type_id"].astype(int)))
    print(f" OK ({time.time() - t0:.1f}s)")

    # â”€â”€ Loop sui periodi: snapshot su DB per ognuno â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    excel_payload = None
    for p in PERIODI:
        df_p, mktp_dfs_p, dash_p = _run_period(
            p, engine, sp_df, spese_gratuite, country_stype, t0
        )
        if p == PERIODO_EXCEL:
            excel_payload = (df_p, mktp_dfs_p, dash_p)

    # â”€â”€ Generazione Excel solo per il periodo principale â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if excel_payload is not None:
        df, mktp_dfs, dash = excel_payload
        global PERIODO
        PERIODO = PERIODO_EXCEL  # garantisce che la nota footer mostri il periodo corretto
        print(f"\nGenerazione Excel (periodo {PERIODO_EXCEL} giorni)...")
        to_excel(df, OUTPUT_FILE, mktp_dfs=mktp_dfs, dashboard_corrieri=dash)
        print(f"File: {OUTPUT_FILE}")
    else:
        print(f"\nATTENZIONE: PERIODO_EXCEL={PERIODO_EXCEL} non presente in PERIODI={PERIODI}, "
              "Excel non generato.")

    elapsed = time.time() - t0
    print(f"\nCompletato in {elapsed:.1f}s")


if __name__ == "__main__":
    main()
