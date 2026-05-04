"""Export dettaglio Danni: rimborso, claims, recovery tramite AEH/allocation."""
import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time, os

PERIODO = 180
DSN = "mysql+pymysql://valerio:v4l3r10!!!@db-master.yeppon.it:3306/smart2"
engine = create_engine(DSN)

t0 = time.time()

# ── 1. Danni RMA base data ──────────────────────────────────────────────
print("1/5 Query Danni RMA base...", end="", flush=True)
sql_danni = f"""
SELECT
  mrp.prodotto       AS id_p,
  p.codice,
  IFNULL(p.nome, '') AS nome_prodotto,
  IFNULL(m.nome, '') AS marca,
  mrp.id_ordine,
  mrp.id_rma,
  mrp.quantita       AS qty_rma,
  IFNULL(mrp.valore_rimborso, 0) AS valore_rimborso,
  IFNULL(mr.valore, 0)           AS valore_claims,
  mr.stato_claims,
  mr.data_creazione  AS data_rma,
  b_agg.costo_unitario,
  b_agg.prezzo_vendita_unit,
  b_agg.mktp         AS marketplace
FROM multi_rma_prodotti mrp
JOIN multi_rma mr        ON mr.id_rma  = mrp.id_rma
JOIN rma_tipo_rma tiporma ON tiporma.id = mr.tipo_rma
LEFT JOIN prodotti p     ON p.id_p     = mrp.prodotto
LEFT JOIN marca m        ON m.id_m     = p.id_m
JOIN (
  SELECT b.id_ordine, b.id_p,
    ROUND(MAX(b.acquistato_tot / b.quantita), 2) AS costo_unitario,
    ROUND(MAX(b.prezzo_vendita / b.quantita), 2) AS prezzo_vendita_unit,
    MAX(b.mktp) AS mktp
  FROM yeppon_stats.bollettato_total b
  JOIN ordini_cliente oc ON oc.id = b.id_ordine AND oc.pixmania NOT IN (10,108)
  WHERE b.prezzo_vendita > 0
    AND b.data_evasione >= CURRENT_DATE - INTERVAL {PERIODO} DAY
  GROUP BY b.id_ordine, b.id_p
) b_agg ON b_agg.id_ordine = mrp.id_ordine AND b_agg.id_p = mrp.prodotto
WHERE tiporma.stato = 'Danni'
  AND mr.data_creazione >= DATE_SUB(CURDATE(), INTERVAL {PERIODO} DAY)
ORDER BY mrp.id_rma, mrp.prodotto
"""
df = pd.read_sql(text(sql_danni), engine)
print(f" {len(df)} righe ({time.time()-t0:.1f}s)")

if df.empty:
    print("Nessun dato Danni trovato.")
    raise SystemExit

# ── 2. Ingressi (LEFT JOIN) ─────────────────────────────────────────────
print("2/5 Query ingressi...", end="", flush=True)
rma_ids = df["id_rma"].unique().tolist()
chunks = [rma_ids[i:i+500] for i in range(0, len(rma_ids), 500)]
parts = []
for chunk in chunks:
    ids_str = ",".join(str(x) for x in chunk)
    q = f"""
    SELECT id_rma_recesso AS id_rma, id_prodotto AS id_p, barcode,
           qta, create_time AS data_ingresso,
           prodotto_integro_sigillato AS integro,
           riparazione, rientro_fornicat AS rientro_forn,
           standby, id_prodotto_outlet AS id_outlet
    FROM ingressi
    WHERE id_rma_recesso IN ({ids_str})
    """
    parts.append(pd.read_sql(text(q), engine))
df_ing = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
print(f" {len(df_ing)} righe ({time.time()-t0:.1f}s)")

# Merge ingressi
if not df_ing.empty:
    # Dedup: keep first ingresso per (id_rma, id_p)
    df_ing = df_ing.sort_values("data_ingresso").drop_duplicates(
        subset=["id_rma", "id_p"], keep="last"
    )
    df = df.merge(df_ing, on=["id_rma", "id_p"], how="left")
else:
    for c in ["barcode", "qta", "data_ingresso", "integro",
              "riparazione", "rientro_forn", "standby", "id_outlet"]:
        df[c] = np.nan

df["ha_ingresso"] = df["barcode"].notna()

# ── 3. Allocation Evasione History (cerca via barcode ingressi → ean) ────
print("3/5 Query allocation_evasione_history...", end="", flush=True)
barcodes_with_ing = df.loc[df["ha_ingresso"], "barcode"].dropna().unique().tolist()
df_aeh = pd.DataFrame()
if barcodes_with_ing:
    aeh_chunks = [barcodes_with_ing[i:i+500] for i in range(0, len(barcodes_with_ing), 500)]
    aeh_parts = []
    for chunk in aeh_chunks:
        bc_str = ",".join(f"'{x}'" for x in chunk)
        q = f"""
        SELECT aeh.ean       AS barcode,
               aeh.id_prodotto AS id_p,
               aeh.id_ordine  AS aeh_id_ordine,
               aeh.id_allocation,
               aeh.tipo       AS aeh_tipo,
               aeh.pkid       AS aeh_pkid,
               aeh.data_aggiornamento AS aeh_data
        FROM allocation_evasione_history aeh
        WHERE aeh.ean IN ({bc_str})
        ORDER BY aeh.ean, aeh.pkid
        """
        aeh_parts.append(pd.read_sql(text(q), engine))
    df_aeh = pd.concat(aeh_parts, ignore_index=True)

if not df_aeh.empty:
    # Keep last row per barcode by pkid
    df_aeh = df_aeh.sort_values("aeh_pkid").drop_duplicates(
        subset=["barcode"], keep="last"
    )
    df_aeh = df_aeh.rename(columns={"id_p": "aeh_id_p"})  # id_p in AEH può differire (es. outlet)
    df = df.merge(df_aeh, on=["barcode"], how="left")
else:
    for c in ["aeh_id_ordine", "id_allocation", "aeh_tipo", "aeh_pkid", "aeh_data"]:
        df[c] = np.nan

print(f" {len(df_aeh)} last-rows ({time.time()-t0:.1f}s)")

# ── 4. Recovery: tipo=S → ordini_cliente + ordini_carrello ──────────────
print("4/5 Lookup rivendita + allocation macrotipo...", end="", flush=True)

# 4a. For tipo=S rows, get order_type from ordini_cliente
mask_s = df["aeh_tipo"] == "S"
df["order_type_rivendita"] = np.nan
df["prezzo_acquisto_recup"] = np.nan

if mask_s.any():
    s_orders = df.loc[mask_s, "aeh_id_ordine"].dropna().astype(int).unique().tolist()
    if s_orders:
        ids_str = ",".join(str(x) for x in s_orders)
        q_oc = f"SELECT id, order_type, pixmania FROM ordini_cliente WHERE id IN ({ids_str})"
        df_oc = pd.read_sql(text(q_oc), engine)
        oc_map = dict(zip(df_oc["id"], df_oc["order_type"]))
        pix_map = dict(zip(df_oc["id"], df_oc["pixmania"]))
        df.loc[mask_s, "order_type_rivendita"] = (
            df.loc[mask_s, "aeh_id_ordine"].map(oc_map)
        )
        df.loc[mask_s, "pixmania_rivendita"] = (
            df.loc[mask_s, "aeh_id_ordine"].map(pix_map)
        )

    # order_type=1/5 OR (order_type=3 AND pixmania=52) → rivenduto
    mask_sold = mask_s & (
        (df["order_type_rivendita"].isin([1, 5]))
        | ((df["order_type_rivendita"] == 3) & (df["pixmania_rivendita"] == 52))
    )
    if mask_sold.any():
        pairs = df.loc[mask_sold, ["aeh_id_ordine", "aeh_id_p"]].dropna().drop_duplicates()
        cond_parts = []
        for _, row in pairs.iterrows():
            cond_parts.append(
                f"(oc2.id_ordine={int(row['aeh_id_ordine'])} "
                f"AND oc2.id_p={int(row['aeh_id_p'])})"
            )
        if cond_parts:
            where_str = " OR ".join(cond_parts)
            q_carr = f"""
            SELECT oc2.id_ordine AS aeh_id_ordine, oc2.id_p AS aeh_id_p,
                   ROUND(oc2.prezzo_acquisto, 2) AS _pa,
                   ROUND(oc2.prezzo * oc2.quantita / (1 + oc2.iva/100), 2) AS _ricavo_netto
            FROM ordini_carrello oc2
            WHERE {where_str}
            """
            df_carr = pd.read_sql(text(q_carr), engine)
            if not df_carr.empty:
                df_carr["aeh_id_ordine"] = df_carr["aeh_id_ordine"].astype(float)
                df_carr["aeh_id_p"] = df_carr["aeh_id_p"].astype(float)
                df = df.merge(df_carr, on=["aeh_id_ordine", "aeh_id_p"],
                              how="left", suffixes=("", "_carr"))
                df["prezzo_acquisto_recup"] = df["prezzo_acquisto_recup"].fillna(df.get("_pa", np.nan))
                df["ricavo_netto_rivendita"] = df.get("_ricavo_netto", np.nan)
                for c in ["_pa", "_ricavo_netto"]:
                    if c in df.columns:
                        df.drop(columns=[c], inplace=True)

# 4b. For tipo=C rows, get allocation macrotipo
mask_c = df["aeh_tipo"] == "C"
df["macrotipo_alloc"] = ""
if mask_c.any():
    alloc_ids = df.loc[mask_c, "id_allocation"].dropna().astype(int).unique().tolist()
    if alloc_ids:
        ids_str = ",".join(str(x) for x in alloc_ids)
        q_alloc = f"""
        SELECT pkid AS id_allocation, macrotipo AS macrotipo_alloc,
               barcode AS barcode_alloc, descrizione AS desc_alloc
        FROM allocation WHERE pkid IN ({ids_str})
        """
        df_alloc = pd.read_sql(text(q_alloc), engine)
        alloc_map = dict(zip(df_alloc["id_allocation"], df_alloc["macrotipo_alloc"]))
        desc_map = dict(zip(df_alloc["id_allocation"], df_alloc["desc_alloc"]))
        barcode_alloc_map = dict(zip(df_alloc["id_allocation"], df_alloc["barcode_alloc"]))
        df.loc[mask_c, "macrotipo_alloc"] = (
            df.loc[mask_c, "id_allocation"].map(alloc_map).fillna("")
        )
        df["desc_alloc"] = ""
        df.loc[mask_c, "desc_alloc"] = (
            df.loc[mask_c, "id_allocation"].map(desc_map).fillna("")
        )
        df["barcode_alloc"] = ""
        df.loc[mask_c, "barcode_alloc"] = (
            df.loc[mask_c, "id_allocation"].map(barcode_alloc_map).fillna("")
        )

print(f" OK ({time.time()-t0:.1f}s)")

# ── 5. Calculate recovery / loss ────────────────────────────────────────
# % di perdita per barcode allocazione (100% = perdita totale, 0% = recupero pieno)
PERDITA_PER_BARCODE = {
    "ING":      0.50,
    "CAT":      0.50,
    "FORNRMA":  0.50,
    "SCARTI":   1.00,
    "DIFPCLI":  1.00,
    "CATCLI":   0.50,
    "RESICOM":  0.00,
    "RESIFOR":  0.00,
    "RESICAT":  0.50,
    "DANNI":    1.00,
    "FOR":      0.00,
    "FORCLI":   0.00,
    "STANDBY":  0.50,
    "RIPAR":    0.50,
    "DGFDANNI": 1.00,
    "TERZISTA": 0.50,
    "CONTES":   0.50,
}

if "barcode_alloc" not in df.columns:
    df["barcode_alloc"] = ""

df["valore_recuperato"] = 0.0
df["perc_perdita"] = np.nan
df["nota_recovery"] = ""

# S + order_type=1/5 OR (order_type=3 AND pixmania=52) → rivenduto
if "pixmania_rivendita" not in df.columns:
    df["pixmania_rivendita"] = np.nan
m1 = (df["aeh_tipo"] == "S") & (
    (df["order_type_rivendita"].isin([1, 5]))
    | ((df["order_type_rivendita"] == 3) & (df["pixmania_rivendita"] == 52))
)
df.loc[m1, "valore_recuperato"] = df.loc[m1, "prezzo_acquisto_recup"].fillna(0)
df.loc[m1, "perc_perdita"] = 0.0
df.loc[m1, "nota_recovery"] = "Rivenduto (ordine " + df.loc[m1, "aeh_id_ordine"].fillna(0).astype(int).astype(str) + ")"

# S + order_type != 1 and != 6 → cerca in allocation_prodotti via barcode(ean)
# Exclude also order_type=3+pixmania=52 (treated as sold above)
m1b = (df["aeh_tipo"] == "S") & (~m1) & (df["order_type_rivendita"] != 6) & df["order_type_rivendita"].notna()
df["ap_macrotipo"] = ""
df["ap_barcode_alloc"] = ""
df["ap_desc_alloc"] = ""

if m1b.any():
    ap_barcodes = df.loc[m1b, "barcode"].dropna().unique().tolist()
    if ap_barcodes:
        ap_chunks = [ap_barcodes[i:i+500] for i in range(0, len(ap_barcodes), 500)]
        ap_parts = []
        for chunk in ap_chunks:
            bc_str = ",".join(f"'{x}'" for x in chunk)
            q_ap = f"""
            SELECT ap.ean AS barcode, a.macrotipo AS ap_macrotipo,
                   a.barcode AS ap_barcode_alloc, a.descrizione AS ap_desc_alloc
            FROM allocation_prodotti ap
            JOIN allocation a ON a.pkid = ap.id_allocation
            WHERE ap.ean IN ({bc_str})
            ORDER BY ap.ean, ap.pkid DESC
            """
            ap_parts.append(pd.read_sql(text(q_ap), engine))
        df_ap = pd.concat(ap_parts, ignore_index=True)
        if not df_ap.empty:
            df_ap = df_ap.drop_duplicates(subset=["barcode"], keep="first")
            ap_macro_map = dict(zip(df_ap["barcode"], df_ap["ap_macrotipo"]))
            ap_bc_map = dict(zip(df_ap["barcode"], df_ap["ap_barcode_alloc"]))
            ap_desc_map = dict(zip(df_ap["barcode"], df_ap["ap_desc_alloc"]))
            df.loc[m1b, "ap_macrotipo"] = df.loc[m1b, "barcode"].map(ap_macro_map).fillna("")
            df.loc[m1b, "ap_barcode_alloc"] = df.loc[m1b, "barcode"].map(ap_bc_map).fillna("")
            df.loc[m1b, "ap_desc_alloc"] = df.loc[m1b, "barcode"].map(ap_desc_map).fillna("")

# S+other con allocation_prodotti → SEDE
m1b_sede = m1b & (df["ap_macrotipo"].str.upper() == "SEDE")
df.loc[m1b_sede, "valore_recuperato"] = df.loc[m1b_sede, "costo_unitario"].fillna(0)
df.loc[m1b_sede, "perc_perdita"] = 0.0
df.loc[m1b_sede, "nota_recovery"] = "In stock SEDE (via alloc_prodotti)"
df.loc[m1b_sede, "macrotipo_alloc"] = "SEDE"

# S+other con allocation_prodotti → ASSISTENZA
m1b_assist = m1b & (df["ap_macrotipo"].str.upper() == "ASSISTENZA")
if m1b_assist.any():
    bc_upper_ap = df.loc[m1b_assist, "ap_barcode_alloc"].str.upper().str.strip()
    pct_loss_ap = bc_upper_ap.map(PERDITA_PER_BARCODE).fillna(0.50)
    df.loc[m1b_assist, "perc_perdita"] = pct_loss_ap.values
    recovery_pct_ap = 1.0 - pct_loss_ap
    df.loc[m1b_assist, "valore_recuperato"] = (
        df.loc[m1b_assist, "costo_unitario"].fillna(0).values * recovery_pct_ap.values
    ).round(2)
    df.loc[m1b_assist, "nota_recovery"] = (
        bc_upper_ap + " (perdita " + (pct_loss_ap * 100).astype(int).astype(str) + "%, via alloc_prodotti)"
    ).values
    df.loc[m1b_assist, "macrotipo_alloc"] = "ASSISTENZA"
    df.loc[m1b_assist, "barcode_alloc"] = df.loc[m1b_assist, "ap_barcode_alloc"]

# S+other con allocation_prodotti → altro macrotipo
m1b_other_ap = m1b & (df["ap_macrotipo"] != "") & ~m1b_sede & ~m1b_assist
df.loc[m1b_other_ap, "nota_recovery"] = "Alloc. " + df.loc[m1b_other_ap, "ap_macrotipo"] + " (via alloc_prodotti)"

# S+other senza allocation_prodotti → nessuna allocazione trovata
m1b_no_ap = m1b & (df["ap_macrotipo"] == "")
df.loc[m1b_no_ap, "nota_recovery"] = "Scaricato (order_type=" + df.loc[m1b_no_ap, "order_type_rivendita"].fillna(0).astype(int).astype(str) + ", no alloc)"

# C + SEDE → back in stock, full recovery
m2 = (df["aeh_tipo"] == "C") & (df["macrotipo_alloc"].str.upper() == "SEDE")
df.loc[m2, "valore_recuperato"] = df.loc[m2, "costo_unitario"].fillna(0)
df.loc[m2, "perc_perdita"] = 0.0
df.loc[m2, "nota_recovery"] = "In stock (SEDE)"

# C + ASSISTENZA → per-barcode loss %
m3 = (df["aeh_tipo"] == "C") & (df["macrotipo_alloc"].str.upper() == "ASSISTENZA")
if m3.any():
    bc_upper = df.loc[m3, "barcode_alloc"].str.upper().str.strip()
    pct_loss = bc_upper.map(PERDITA_PER_BARCODE).fillna(0.50)  # default 50% if unknown
    df.loc[m3, "perc_perdita"] = pct_loss.values
    recovery_pct = 1.0 - pct_loss
    df.loc[m3, "valore_recuperato"] = (
        df.loc[m3, "costo_unitario"].fillna(0).values * recovery_pct.values
    ).round(2)
    df.loc[m3, "nota_recovery"] = (
        bc_upper + " (perdita " + (pct_loss * 100).astype(int).astype(str) + "%)"
    ).values

# C + other macrotipo
m4 = (df["aeh_tipo"] == "C") & ~m2 & ~m3 & df["macrotipo_alloc"].notna() & (df["macrotipo_alloc"] != "")
df.loc[m4, "nota_recovery"] = "Alloc. " + df.loc[m4, "macrotipo_alloc"]

# No ingresso
m_no_ing = ~df["ha_ingresso"]
df.loc[m_no_ing, "nota_recovery"] = "Nessun ingresso"

# Has ingresso but no AEH
m_no_aeh = df["ha_ingresso"] & df["aeh_tipo"].isna()
df.loc[m_no_aeh, "nota_recovery"] = (
    "Ingresso OK, no AEH (bc=" + df.loc[m_no_aeh, "barcode"].fillna("").astype(str) + ")"
)

# Flag righe dove rimborso era 0 (perdita piena al costo acquisto, no proporzione)
rimborso_zero = df["valore_rimborso"] == 0

# Se valore_rimborso è 0, usa il costo di acquisto del prodotto
df.loc[rimborso_zero, "valore_rimborso"] = df.loc[
    rimborso_zero, "costo_unitario"
].fillna(0)

# Claims 0.01 → 0 (placeholder, non reale)
df.loc[df["valore_claims"] == 0.01, "valore_claims"] = 0

# Costo effettivo perdita: (valore_rimborso / prezzo_vendita) * costo_unitario
# = percentuale rimborsata applicata al costo di acquisto
df["perc_rimborso"] = (
    df["valore_rimborso"] / df["prezzo_vendita_unit"].replace(0, np.nan)
).clip(upper=1.0).fillna(1.0)
df["costo_perdita"] = (df["perc_rimborso"] * df["costo_unitario"].fillna(0)).round(2)

# Override: dove rimborso era 0, costo_perdita = costo_unitario (100%) senza proporzione
df.loc[rimborso_zero, "perc_rimborso"] = 1.0
df.loc[rimborso_zero, "costo_perdita"] = df.loc[rimborso_zero, "costo_unitario"].fillna(0).round(2)

# Rivenduti (S+1): valore recuperato = ricavo netto rivendita (prezzo*qty / (1+iva%))
if "ricavo_netto_rivendita" not in df.columns:
    df["ricavo_netto_rivendita"] = np.nan
m_rivend = m1  # same mask: order_type=1 OR (3+pix52)
df.loc[m_rivend, "valore_recuperato"] = df.loc[m_rivend, "ricavo_netto_rivendita"].fillna(0)

# S + order_type=6 → scarico interno, valore recuperato = costo perdita → perdita netta 0
m_ot6 = (df["aeh_tipo"] == "S") & (df["order_type_rivendita"] == 6) & ~m1
df.loc[m_ot6, "valore_recuperato"] = df.loc[m_ot6, "costo_perdita"]
df.loc[m_ot6, "perc_perdita"] = 0.0
df.loc[m_ot6, "nota_recovery"] = "Scarico interno (order_type=6)"

# Perdita netta = costo_perdita - claims ricevuti - valore recuperato
df["perdita_netta"] = (
    df["costo_perdita"] - df["valore_claims"] - df["valore_recuperato"]
).round(2)

# ── 6. Excel output ─────────────────────────────────────────────────────
print("5/5 Generazione Excel...", end="", flush=True)

COLS = [
    ("id_p",                 "ID Prodotto",      12),
    ("codice",               "Codice",           16),
    ("nome_prodotto",        "Nome",             40),
    ("marca",                "Marca",            16),
    ("id_ordine",            "ID Ordine Orig.",  14),
    ("id_rma",               "ID RMA",           10),
    ("data_rma",             "Data RMA",         12),
    ("qty_rma",              "Qty",               6),
    ("marketplace",          "Marketplace",      14),
    ("prezzo_vendita_unit",  "Prezzo Vendita €", 14),
    ("costo_unitario",       "Costo Acq. €",    12),
    ("valore_rimborso",      "Rimborso Cliente €",16),
    ("perc_rimborso",        "% Rimborso",       10),
    ("costo_perdita",        "Costo Perdita €",  14),
    ("valore_claims",        "Claims Corriere €", 16),
    ("ha_ingresso",          "Ingresso?",        10),
    ("data_ingresso",        "Data Ingresso",    12),
    ("aeh_tipo",             "Ultimo Mov.",      10),
    ("aeh_id_ordine",        "Ordine Rivend.",   14),
    ("macrotipo_alloc",      "Macrotipo Alloc.", 14),
    ("barcode_alloc",        "Barcode Alloc.",   14),
    ("desc_alloc",           "Allocazione",      14),
    ("perc_perdita",         "% Perdita",        10),
    ("valore_recuperato",    "Valore Recuperato €",16),
    ("nota_recovery",        "Nota Recovery",    28),
    ("perdita_netta",        "Perdita Netta €",  14),
]

fields  = [c[0] for c in COLS]
headers = [c[1] for c in COLS]
widths  = [c[2] for c in COLS]

# Ensure all fields exist
for f in fields:
    if f not in df.columns:
        df[f] = ""

out = df[fields].copy()
out["ha_ingresso"] = out["ha_ingresso"].map({True: "Sì", False: "No"})
out.columns = headers

out_path = os.path.join(
    os.path.expanduser("~"), "Downloads",
    f"dettaglio_danni_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
)

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    out.to_excel(writer, sheet_name="Danni", index=False)
    ws = writer.sheets["Danni"]

    # Styles
    F_HDR  = PatternFill("solid", fgColor="1F3864")
    T_HDR  = Font(bold=True, color="FFFFFF", size=10)
    F_SI   = PatternFill("solid", fgColor="C6EFCE")  # green
    F_NO   = PatternFill("solid", fgColor="FFC7CE")  # red
    F_GAIN = PatternFill("solid", fgColor="DDEBF7")  # light blue
    F_LOSS = PatternFill("solid", fgColor="FCE4D6")  # orange
    THIN   = Side(style="thin", color="D0D0D0")
    BORDER = Border(bottom=THIN)

    # Header row
    ws.row_dimensions[1].height = 34
    for ci in range(1, len(COLS) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = F_HDR
        c.font = T_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]

    EUR_FMT = '#,##0.00 €'
    eur_cols = ["prezzo_vendita_unit", "costo_unitario", "valore_rimborso",
                "costo_perdita", "valore_claims", "valore_recuperato", "perdita_netta"]
    pct_cols = ["perc_rimborso", "perc_perdita"]

    ing_col    = fields.index("ha_ingresso") + 1
    perd_col   = fields.index("perdita_netta") + 1
    recup_col  = fields.index("valore_recuperato") + 1
    perc_col   = fields.index("perc_perdita") + 1

    for ri in range(2, len(out) + 2):
        for col_name in eur_cols:
            ci = fields.index(col_name) + 1
            ws.cell(row=ri, column=ci).number_format = EUR_FMT

        # Format: percentuali
        for col_name in pct_cols:
            ci = fields.index(col_name) + 1
            ws.cell(row=ri, column=ci).number_format = '0%'

        # Color: Ha Ingresso
        cell_ing = ws.cell(row=ri, column=ing_col)
        if cell_ing.value == "Sì":
            cell_ing.fill = F_SI; cell_ing.font = Font(color="006100")
        else:
            cell_ing.fill = F_NO; cell_ing.font = Font(color="9C0006")

        # Color: Perdita netta
        v = ws.cell(row=ri, column=perd_col).value
        if v is not None:
            if v > 0:
                ws.cell(row=ri, column=perd_col).fill = F_LOSS
                ws.cell(row=ri, column=perd_col).font = Font(color="C00000", bold=True)
            elif v < 0:
                ws.cell(row=ri, column=perd_col).fill = F_GAIN
                ws.cell(row=ri, column=perd_col).font = Font(color="006100")

        # Color: Valore recuperato > 0
        vr = ws.cell(row=ri, column=recup_col).value
        if vr and vr > 0:
            ws.cell(row=ri, column=recup_col).fill = F_GAIN

    # Summary row
    last_row = len(out) + 2
    ws.cell(row=last_row, column=1).value = "TOTALE"
    ws.cell(row=last_row, column=1).font = Font(bold=True, size=11)
    for col_name in eur_cols:
        ci = fields.index(col_name) + 1
        col_letter = get_column_letter(ci)
        ws.cell(row=last_row, column=ci).value = (
            f"=SUM({col_letter}2:{col_letter}{last_row-1})"
        )
        ws.cell(row=last_row, column=ci).number_format = EUR_FMT
        ws.cell(row=last_row, column=ci).font = Font(bold=True, size=11)

    # Freeze panes
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(out)+1}"

print(f" OK ({time.time()-t0:.1f}s)")

# Quick stats
n_with_ing = df["ha_ingresso"].sum()
n_with_aeh = df["aeh_tipo"].notna().sum()
n_sold     = ((df["aeh_tipo"] == "S") & (df["order_type_rivendita"] == 1)).sum()
n_sede     = ((df["aeh_tipo"] == "C") & (df["macrotipo_alloc"].str.upper() == "SEDE")).sum()
n_assist   = ((df["aeh_tipo"] == "C") & (df["macrotipo_alloc"].str.upper() == "ASSISTENZA")).sum()
tot_rimb   = df["valore_rimborso"].sum()
tot_cperd  = df["costo_perdita"].sum()
tot_claims = df["valore_claims"].sum()
tot_recup  = df["valore_recuperato"].sum()
tot_perd   = df["perdita_netta"].sum()

print(f"\n--- Riepilogo Danni ({len(df)} righe) ---")
print(f"  Con ingresso:    {n_with_ing:>4}")
print(f"  Con AEH:         {n_with_aeh:>4}")
print(f"  Rivenduti (S+1): {n_sold:>4}")
print(f"  In stock (SEDE): {n_sede:>4}")
print(f"  Assistenza:      {n_assist:>4}")
print(f"  Tot. Rimborso:   {tot_rimb:>12,.2f} €")
print(f"  Tot. Costo Perd.:{tot_cperd:>12,.2f} €")
print(f"  Tot. Claims:     {tot_claims:>12,.2f} €")
print(f"  Tot. Recuperato: {tot_recup:>12,.2f} €")
print(f"  Tot. Perdita:    {tot_perd:>12,.2f} €")
print(f"\nFile: {out_path}")
