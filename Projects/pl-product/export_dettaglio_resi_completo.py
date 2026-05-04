"""Export dettaglio TUTTI i resi: rimborso, claims, recovery tramite AEH/allocation.
Un foglio Excel per ogni tipologia RMA + foglio Riepilogo."""
import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time, os

PERIODO = 180
DSN = "mysql+pymysql://valerio:v4l3r10!!!@db-master.yeppon.it:3306/smart2"
engine = create_engine(DSN)

t0 = time.time()

# ── 1. ALL RMA base data ────────────────────────────────────────────────
print("1/5 Query RMA base (tutte le tipologie)...", end="", flush=True)
sql_all = f"""
SELECT
  mrp.prodotto       AS id_p,
  p.codice,
  IFNULL(p.nome, '') AS nome_prodotto,
  IFNULL(m.nome, '') AS marca,
  mrp.id_ordine,
  mrp.id_rma,
  tiporma.stato      AS tipo_rma,
  mrp.quantita       AS qty_rma,
  IFNULL(mrp.valore_rimborso, 0) AS valore_rimborso,
  IFNULL(mr.valore, 0)           AS valore_claims,
  mr.stato_claims,
  mr.data_creazione  AS data_rma,
  b_agg.costo_unitario,
  b_agg.prezzo_vendita_unit,
  b_agg.mktp         AS marketplace
FROM multi_rma_prodotti mrp
JOIN multi_rma mr        ON mr.id_rma  = mrp.id_rma
JOIN rma_tipo_rma tiporma ON tiporma.id = mr.tipo_rma
LEFT JOIN prodotti p     ON p.id_p     = mrp.prodotto
LEFT JOIN marca m        ON m.id_m     = p.id_m
JOIN (
  SELECT b.id_ordine, b.id_p,
    ROUND(MAX(b.acquistato_tot / b.quantita), 2) AS costo_unitario,
    ROUND(MAX(b.prezzo_vendita / b.quantita), 2) AS prezzo_vendita_unit,
    MAX(b.mktp) AS mktp
  FROM yeppon_stats.bollettato_total b
  JOIN ordini_cliente oc ON oc.id = b.id_ordine AND oc.pixmania NOT IN (10,108)
  WHERE b.prezzo_vendita > 0
    AND b.data_evasione >= CURRENT_DATE - INTERVAL {PERIODO} DAY
  GROUP BY b.id_ordine, b.id_p
) b_agg ON b_agg.id_ordine = mrp.id_ordine AND b_agg.id_p = mrp.prodotto
WHERE mr.data_creazione >= DATE_SUB(CURDATE(), INTERVAL {PERIODO} DAY)
ORDER BY tiporma.stato, mrp.id_rma, mrp.prodotto
"""
df = pd.read_sql(text(sql_all), engine)
print(f" {len(df)} righe ({time.time()-t0:.1f}s)")

if df.empty:
    print("Nessun dato trovato.")
    raise SystemExit

# ── 2. Ingressi (LEFT JOIN) ─────────────────────────────────────────────
print("2/5 Query ingressi...", end="", flush=True)
rma_ids = df["id_rma"].unique().tolist()
chunks = [rma_ids[i:i+500] for i in range(0, len(rma_ids), 500)]
parts = []
for chunk in chunks:
    ids_str = ",".join(str(x) for x in chunk)
    q = f"""
    SELECT id_rma_recesso AS id_rma, id_prodotto AS id_p, barcode,
           qta, create_time AS data_ingresso,
           prodotto_integro_sigillato AS integro,
           riparazione, rientro_fornicat AS rientro_forn,
           standby, id_prodotto_outlet AS id_outlet
    FROM ingressi
    WHERE id_rma_recesso IN ({ids_str})
    """
    parts.append(pd.read_sql(text(q), engine))
df_ing = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
print(f" {len(df_ing)} righe ({time.time()-t0:.1f}s)")

if not df_ing.empty:
    df_ing = df_ing.sort_values("data_ingresso").drop_duplicates(
        subset=["id_rma", "id_p"], keep="last"
    )
    df = df.merge(df_ing, on=["id_rma", "id_p"], how="left")
else:
    for c in ["barcode", "qta", "data_ingresso", "integro",
              "riparazione", "rientro_forn", "standby", "id_outlet"]:
        df[c] = np.nan

df["ha_ingresso"] = df["barcode"].notna()

# ── 3. Allocation Evasione History (cerca via barcode ingressi → ean) ────
print("3/5 Query allocation_evasione_history...", end="", flush=True)
barcodes_with_ing = df.loc[df["ha_ingresso"], "barcode"].dropna().unique().tolist()
df_aeh = pd.DataFrame()
if barcodes_with_ing:
    aeh_chunks = [barcodes_with_ing[i:i+500] for i in range(0, len(barcodes_with_ing), 500)]
    aeh_parts = []
    for chunk in aeh_chunks:
        bc_str = ",".join(f"'{x}'" for x in chunk)
        q = f"""
        SELECT aeh.ean       AS barcode,
               aeh.id_prodotto AS id_p,
               aeh.id_ordine  AS aeh_id_ordine,
               aeh.id_allocation,
               aeh.tipo       AS aeh_tipo,
               aeh.pkid       AS aeh_pkid,
               aeh.data_aggiornamento AS aeh_data
        FROM allocation_evasione_history aeh
        WHERE aeh.ean IN ({bc_str})
        ORDER BY aeh.ean, aeh.pkid
        """
        aeh_parts.append(pd.read_sql(text(q), engine))
    df_aeh = pd.concat(aeh_parts, ignore_index=True)

if not df_aeh.empty:
    df_aeh = df_aeh.sort_values("aeh_pkid").drop_duplicates(
        subset=["barcode"], keep="last"
    )
    df_aeh = df_aeh.rename(columns={"id_p": "aeh_id_p"})
    df = df.merge(df_aeh, on=["barcode"], how="left")
else:
    for c in ["aeh_id_ordine", "id_allocation", "aeh_tipo", "aeh_pkid", "aeh_data", "aeh_id_p"]:
        df[c] = np.nan

print(f" {len(df_aeh)} last-rows ({time.time()-t0:.1f}s)")

# ── 4. Recovery: tipo=S → ordini_cliente + ordini_carrello ──────────────
print("4/5 Lookup rivendita + allocation macrotipo...", end="", flush=True)

mask_s = df["aeh_tipo"] == "S"
df["order_type_rivendita"] = np.nan
df["prezzo_acquisto_recup"] = np.nan

if mask_s.any():
    s_orders = df.loc[mask_s, "aeh_id_ordine"].dropna().astype(int).unique().tolist()
    if s_orders:
        ids_str = ",".join(str(x) for x in s_orders)
        q_oc = f"SELECT id, order_type, pixmania FROM ordini_cliente WHERE id IN ({ids_str})"
        df_oc = pd.read_sql(text(q_oc), engine)
        oc_map = dict(zip(df_oc["id"], df_oc["order_type"]))
        pix_map = dict(zip(df_oc["id"], df_oc["pixmania"]))
        df.loc[mask_s, "order_type_rivendita"] = (
            df.loc[mask_s, "aeh_id_ordine"].map(oc_map)
        )
        df.loc[mask_s, "pixmania_rivendita"] = (
            df.loc[mask_s, "aeh_id_ordine"].map(pix_map)
        )

    # order_type=1/5 OR (order_type=3 AND pixmania=52) → rivenduto
    mask_sold = mask_s & (
        (df["order_type_rivendita"].isin([1, 5]))
        | ((df["order_type_rivendita"] == 3) & (df["pixmania_rivendita"] == 52))
    )
    if mask_sold.any():
        pairs = df.loc[mask_sold, ["aeh_id_ordine", "aeh_id_p"]].dropna().drop_duplicates()
        cond_parts = []
        for _, row in pairs.iterrows():
            cond_parts.append(
                f"(oc2.id_ordine={int(row['aeh_id_ordine'])} "
                f"AND oc2.id_p={int(row['aeh_id_p'])})"
            )
        if cond_parts:
            where_str = " OR ".join(cond_parts)
            # Nota: la quantita usata per il ricavo netto e' il MINORE tra
            # la quantita dell'ordine di rivendita (oc2.quantita) e la
            # quantita resa (qty_rma). Es.: se sono stati rivenduti 2 pezzi
            # ma il reso ne contiene 5, il valore recuperato e' di 2 pezzi;
            # viceversa se nell'ordine ce ne sono 8 ma resi solo 1, il
            # valore recuperato e' di 1 pezzo.
            q_carr = f"""
            SELECT oc2.id_ordine AS aeh_id_ordine, oc2.id_p AS aeh_id_p,
                   ROUND(oc2.prezzo_acquisto, 2) AS _pa,
                   ROUND(oc2.prezzo / (1 + oc2.iva/100), 4) AS _prezzo_unit_netto,
                   oc2.quantita AS _qty_rivendita
            FROM ordini_carrello oc2
            WHERE {where_str}
            """
            df_carr = pd.read_sql(text(q_carr), engine)
            if not df_carr.empty:
                df_carr["aeh_id_ordine"] = df_carr["aeh_id_ordine"].astype(float)
                df_carr["aeh_id_p"] = df_carr["aeh_id_p"].astype(float)
                df = df.merge(df_carr, on=["aeh_id_ordine", "aeh_id_p"],
                              how="left", suffixes=("", "_carr"))
                df["prezzo_acquisto_recup"] = df["prezzo_acquisto_recup"].fillna(df.get("_pa", np.nan))
                # qty effettiva = min(qty_rivendita, qty_rma)
                qty_effettiva = np.minimum(
                    df.get("_qty_rivendita", pd.Series(np.nan, index=df.index)).fillna(0),
                    df["qty_rma"].fillna(0),
                )
                df["ricavo_netto_rivendita"] = (
                    df.get("_prezzo_unit_netto", np.nan) * qty_effettiva
                ).round(2)
                for c in ["_pa", "_prezzo_unit_netto", "_qty_rivendita"]:
                    if c in df.columns:
                        df.drop(columns=[c], inplace=True)

# 4b. For tipo=C rows, get allocation macrotipo
mask_c = df["aeh_tipo"] == "C"
df["macrotipo_alloc"] = ""
if mask_c.any():
    alloc_ids = df.loc[mask_c, "id_allocation"].dropna().astype(int).unique().tolist()
    if alloc_ids:
        ids_str = ",".join(str(x) for x in alloc_ids)
        q_alloc = f"""
        SELECT pkid AS id_allocation, macrotipo AS macrotipo_alloc,
               barcode AS barcode_alloc, descrizione AS desc_alloc
        FROM allocation WHERE pkid IN ({ids_str})
        """
        df_alloc = pd.read_sql(text(q_alloc), engine)
        alloc_map = dict(zip(df_alloc["id_allocation"], df_alloc["macrotipo_alloc"]))
        desc_map = dict(zip(df_alloc["id_allocation"], df_alloc["desc_alloc"]))
        barcode_alloc_map = dict(zip(df_alloc["id_allocation"], df_alloc["barcode_alloc"]))
        df.loc[mask_c, "macrotipo_alloc"] = (
            df.loc[mask_c, "id_allocation"].map(alloc_map).fillna("")
        )
        df["desc_alloc"] = ""
        df.loc[mask_c, "desc_alloc"] = (
            df.loc[mask_c, "id_allocation"].map(desc_map).fillna("")
        )
        df["barcode_alloc"] = ""
        df.loc[mask_c, "barcode_alloc"] = (
            df.loc[mask_c, "id_allocation"].map(barcode_alloc_map).fillna("")
        )

print(f" OK ({time.time()-t0:.1f}s)")

# ── 5. Calculate recovery / loss ────────────────────────────────────────
PERDITA_PER_BARCODE = {
    "ING":      0.50,
    "CAT":      0.50,
    "FORNRMA":  0.50,
    "SCARTI":   1.00,
    "DIFPCLI":  1.00,
    "CATCLI":   0.50,
    "RESICOM":  0.00,
    "RESIFOR":  0.00,
    "RESICAT":  0.50,
    "DANNI":    1.00,
    "FOR":      0.00,
    "FORCLI":   0.00,
    "STANDBY":  0.50,
    "RIPAR":    0.50,
    "DGFDANNI": 1.00,
    "TERZISTA": 0.50,
    "CONTES":   0.50,
}

if "barcode_alloc" not in df.columns:
    df["barcode_alloc"] = ""

df["valore_recuperato"] = 0.0
df["perc_perdita"] = np.nan
df["nota_recovery"] = ""

# S + order_type=1/5 OR (order_type=3 AND pixmania=52) → rivenduto
if "pixmania_rivendita" not in df.columns:
    df["pixmania_rivendita"] = np.nan
m1 = (df["aeh_tipo"] == "S") & (
    (df["order_type_rivendita"].isin([1, 5]))
    | ((df["order_type_rivendita"] == 3) & (df["pixmania_rivendita"] == 52))
)
df.loc[m1, "perc_perdita"] = 0.0
df.loc[m1, "nota_recovery"] = "Rivenduto (ordine " + df.loc[m1, "aeh_id_ordine"].fillna(0).astype(int).astype(str) + ")"

# S + order_type != 1 and != 6 → cerca in allocation_prodotti via barcode(ean)
# Exclude also order_type=3+pixmania=52 (treated as sold above)
m1b = (df["aeh_tipo"] == "S") & (~m1) & (df["order_type_rivendita"] != 6) & df["order_type_rivendita"].notna()
df["ap_macrotipo"] = ""
df["ap_barcode_alloc"] = ""
df["ap_desc_alloc"] = ""

if m1b.any():
    ap_barcodes = df.loc[m1b, "barcode"].dropna().unique().tolist()
    if ap_barcodes:
        ap_chunks = [ap_barcodes[i:i+500] for i in range(0, len(ap_barcodes), 500)]
        ap_parts = []
        for chunk in ap_chunks:
            bc_str = ",".join(f"'{x}'" for x in chunk)
            q_ap = f"""
            SELECT ap.ean AS barcode, a.macrotipo AS ap_macrotipo,
                   a.barcode AS ap_barcode_alloc, a.descrizione AS ap_desc_alloc
            FROM allocation_prodotti ap
            JOIN allocation a ON a.pkid = ap.id_allocation
            WHERE ap.ean IN ({bc_str})
            ORDER BY ap.ean, ap.pkid DESC
            """
            ap_parts.append(pd.read_sql(text(q_ap), engine))
        df_ap = pd.concat(ap_parts, ignore_index=True)
        if not df_ap.empty:
            # keep last allocation per barcode
            df_ap = df_ap.drop_duplicates(subset=["barcode"], keep="first")
            ap_macro_map = dict(zip(df_ap["barcode"], df_ap["ap_macrotipo"]))
            ap_bc_map = dict(zip(df_ap["barcode"], df_ap["ap_barcode_alloc"]))
            ap_desc_map = dict(zip(df_ap["barcode"], df_ap["ap_desc_alloc"]))
            df.loc[m1b, "ap_macrotipo"] = df.loc[m1b, "barcode"].map(ap_macro_map).fillna("")
            df.loc[m1b, "ap_barcode_alloc"] = df.loc[m1b, "barcode"].map(ap_bc_map).fillna("")
            df.loc[m1b, "ap_desc_alloc"] = df.loc[m1b, "barcode"].map(ap_desc_map).fillna("")

# S+other con allocation_prodotti → SEDE
m1b_sede = m1b & (df["ap_macrotipo"].str.upper() == "SEDE")
df.loc[m1b_sede, "valore_recuperato"] = df.loc[m1b_sede, "costo_unitario"].fillna(0)
df.loc[m1b_sede, "perc_perdita"] = 0.0
df.loc[m1b_sede, "nota_recovery"] = "In stock SEDE (via alloc_prodotti)"
df.loc[m1b_sede, "macrotipo_alloc"] = "SEDE"

# S+other con allocation_prodotti → ASSISTENZA
m1b_assist = m1b & (df["ap_macrotipo"].str.upper() == "ASSISTENZA")
if m1b_assist.any():
    bc_upper_ap = df.loc[m1b_assist, "ap_barcode_alloc"].str.upper().str.strip()
    pct_loss_ap = bc_upper_ap.map(PERDITA_PER_BARCODE).fillna(0.50)
    df.loc[m1b_assist, "perc_perdita"] = pct_loss_ap.values
    recovery_pct_ap = 1.0 - pct_loss_ap
    df.loc[m1b_assist, "valore_recuperato"] = (
        df.loc[m1b_assist, "costo_unitario"].fillna(0).values * recovery_pct_ap.values
    ).round(2)
    df.loc[m1b_assist, "nota_recovery"] = (
        bc_upper_ap + " (perdita " + (pct_loss_ap * 100).astype(int).astype(str) + "%, via alloc_prodotti)"
    ).values
    df.loc[m1b_assist, "macrotipo_alloc"] = "ASSISTENZA"
    df.loc[m1b_assist, "barcode_alloc"] = df.loc[m1b_assist, "ap_barcode_alloc"]

# S+other con allocation_prodotti → altro macrotipo
m1b_other_ap = m1b & (df["ap_macrotipo"] != "") & ~m1b_sede & ~m1b_assist
df.loc[m1b_other_ap, "nota_recovery"] = "Alloc. " + df.loc[m1b_other_ap, "ap_macrotipo"] + " (via alloc_prodotti)"

# S+other senza allocation_prodotti → nessuns allocazione trovata
m1b_no_ap = m1b & (df["ap_macrotipo"] == "")
df.loc[m1b_no_ap, "nota_recovery"] = "Scaricato (order_type=" + df.loc[m1b_no_ap, "order_type_rivendita"].fillna(0).astype(int).astype(str) + ", no alloc)"

# C + SEDE → in stock, full recovery
m2 = (df["aeh_tipo"] == "C") & (df["macrotipo_alloc"].str.upper() == "SEDE")
df.loc[m2, "valore_recuperato"] = df.loc[m2, "costo_unitario"].fillna(0)
df.loc[m2, "perc_perdita"] = 0.0
df.loc[m2, "nota_recovery"] = "In stock (SEDE)"

# C + ASSISTENZA → per-barcode loss %
m3 = (df["aeh_tipo"] == "C") & (df["macrotipo_alloc"].str.upper() == "ASSISTENZA")
if m3.any():
    bc_upper = df.loc[m3, "barcode_alloc"].str.upper().str.strip()
    pct_loss = bc_upper.map(PERDITA_PER_BARCODE).fillna(0.50)
    df.loc[m3, "perc_perdita"] = pct_loss.values
    recovery_pct = 1.0 - pct_loss
    df.loc[m3, "valore_recuperato"] = (
        df.loc[m3, "costo_unitario"].fillna(0).values * recovery_pct.values
    ).round(2)
    df.loc[m3, "nota_recovery"] = (
        bc_upper + " (perdita " + (pct_loss * 100).astype(int).astype(str) + "%)"
    ).values

# C + other macrotipo
m4 = (df["aeh_tipo"] == "C") & ~m2 & ~m3 & df["macrotipo_alloc"].notna() & (df["macrotipo_alloc"] != "")
df.loc[m4, "nota_recovery"] = "Alloc. " + df.loc[m4, "macrotipo_alloc"]

# No ingresso
m_no_ing = ~df["ha_ingresso"]
df.loc[m_no_ing, "nota_recovery"] = "Nessun ingresso"

# Has ingresso but no AEH
m_no_aeh = df["ha_ingresso"] & df["aeh_tipo"].isna()
df.loc[m_no_aeh, "nota_recovery"] = (
    "Ingresso OK, no AEH (bc=" + df.loc[m_no_aeh, "barcode"].fillna("").astype(str) + ")"
)

# Mancato Ritiro: il prodotto viene sempre riallocato → recupero 100% al costo acquisto
m_mr = df["tipo_rma"] == "Mancato Ritiro"
df.loc[m_mr, "valore_recuperato"] = df.loc[m_mr, "costo_unitario"].fillna(0)
df.loc[m_mr, "perc_perdita"] = 0.0
df.loc[m_mr, "nota_recovery"] = "Riallocato (Mancato Ritiro)"

# Giacenza: valore_recuperato = costo_perdita → perdita netta sempre 0
# (spostato dopo calcolo costo_perdita)

# ── 4c. Costo spedizione outbound + pixmania ────────────────────────────
# Per Recesso/Giacenza/Danni (escluso Stockly = pixmania 52) aggiungiamo
# il costo della spedizione di andata al costo_perdita (assumiamo costo
# di rientro ~= costo outbound). Il costo va aggiunto UNA SOLA VOLTA per
# id_rma (è dell'ordine intero, non per prodotto): viene assegnato alla
# prima riga di ogni id_rma, le altre restano a 0.
print("4c/5 Lookup costo spedizione outbound...", end="", flush=True)
df["costo_sped_rientro"] = 0.0
df["pixmania_ordine"] = pd.NA

_ord_ids = df["id_ordine"].dropna().astype(int).unique().tolist()
if _ord_ids:
    _pix_parts = []
    _sped_parts = []
    for _chunk in [_ord_ids[i:i+500] for i in range(0, len(_ord_ids), 500)]:
        _ids_str = ",".join(str(x) for x in _chunk)
        _pix_parts.append(pd.read_sql(text(
            f"SELECT id AS id_ordine, pixmania FROM ordini_cliente WHERE id IN ({_ids_str})"
        ), engine))
        _sped_parts.append(pd.read_sql(text(f"""
            SELECT b.id_ordine,
                   IFNULL(MAX(b.spese_sped_corriere), 0) AS costo_sped_corriere,
                   IFNULL(MAX(b.spese_sped_grat + b.spese_sped_cliente), 0) AS sped_incasso
            FROM yeppon_stats.bollettato_total b
            WHERE b.id_ordine IN ({_ids_str})
            GROUP BY b.id_ordine
        """), engine))
    _df_pix = pd.concat(_pix_parts, ignore_index=True)
    _df_sped = pd.concat(_sped_parts, ignore_index=True)

    # Fallback: usa il costo corriere se > 0, altrimenti l'incasso spedizione
    _df_sped["costo_sped_outbound"] = _df_sped["costo_sped_corriere"].where(
        _df_sped["costo_sped_corriere"] > 0, _df_sped["sped_incasso"]
    )

    _pix_map = dict(zip(_df_pix["id_ordine"], _df_pix["pixmania"]))
    _sped_map = dict(zip(_df_sped["id_ordine"], _df_sped["costo_sped_outbound"]))

    df["pixmania_ordine"] = df["id_ordine"].map(_pix_map)
    df["_sped_outbound"] = df["id_ordine"].map(_sped_map).fillna(0.0).astype(float)

    # Maschera: Recesso/Giacenza/Danni/Prodotto non conforme AND non Stockly (pixmania != 52)
    # AND riga con ingresso (se nessuna riga dell'RMA ha ingresso, costo = 0)
    _m_add_sped = (
        df["tipo_rma"].isin(["Recesso", "Giacenza", "Danni", "Prodotto non conforme"])
        & (df["pixmania_ordine"] != 52)
        & df["ha_ingresso"]
    )

    # Una sola volta per id_rma (assegna alla PRIMA riga di ogni id_rma)
    if _m_add_sped.any():
        _first_idx = (
            df.loc[_m_add_sped]
              .groupby("id_rma", as_index=False)
              .head(1)
              .index
        )
        df.loc[_first_idx, "costo_sped_rientro"] = (
            df.loc[_first_idx, "_sped_outbound"].round(2)
        )
    df.drop(columns=["_sped_outbound"], inplace=True)

print(f" OK ({time.time()-t0:.1f}s)")

# Flag righe dove rimborso era 0 (perdita piena al costo acquisto, no proporzione)
rimborso_zero = df["valore_rimborso"] == 0

# Se valore_rimborso è 0, usa il costo di acquisto come rimborso visualizzato
df.loc[rimborso_zero, "valore_rimborso"] = df.loc[
    rimborso_zero, "costo_unitario"
].fillna(0)

# Claims 0.01 → 0
df.loc[df["valore_claims"] == 0.01, "valore_claims"] = 0

# Costo perdita:
# - righe con rimborso originale > 0 → proporzionale: (rimborso / prezzo_vendita) * costo_acquisto
# - righe con rimborso originale = 0 → perdita piena: costo_acquisto (no proporzione)
df["perc_rimborso"] = (
    df["valore_rimborso"] / df["prezzo_vendita_unit"].replace(0, np.nan)
).clip(upper=1.0).fillna(1.0)
df["costo_perdita"] = (df["perc_rimborso"] * df["costo_unitario"].fillna(0)).round(2)

# Override: dove rimborso era 0, costo_perdita = costo_unitario (100%) senza proporzione
df.loc[rimborso_zero, "perc_rimborso"] = 1.0
df.loc[rimborso_zero, "costo_perdita"] = df.loc[rimborso_zero, "costo_unitario"].fillna(0).round(2)

# Nota: il costo_sped_rientro NON viene sommato a costo_perdita (resta
# separato come voce a sé), ma viene aggiunto direttamente in perdita_netta
# più sotto.

# Rivenduti (S+1): valore recuperato = min(ricavo netto rivendita, costo perdita)
# Cap: il valore recuperato non puo' superare il costo della perdita.
# Cosi' il differenziale (perdita_netta = costo_perdita - claims - recuperato)
# e' al massimo 0 quando rivendiamo a un prezzo >= della perdita
# (nessun "guadagno fittizio" dalla rivendita di un reso), mentre resta
# positivo (perdita reale) quando il prezzo di rivendita e' inferiore
# al costo della perdita.
if "ricavo_netto_rivendita" not in df.columns:
    df["ricavo_netto_rivendita"] = np.nan
m_rivend = m1  # same mask: order_type=1 OR (3+pix52)
df.loc[m_rivend, "valore_recuperato"] = np.minimum(
    df.loc[m_rivend, "ricavo_netto_rivendita"].fillna(0).values,
    df.loc[m_rivend, "costo_perdita"].fillna(0).values,
)

# S + order_type=6 → scarico interno, valore recuperato = costo perdita → perdita netta 0
m_ot6 = (df["aeh_tipo"] == "S") & (df["order_type_rivendita"] == 6) & ~m1
df.loc[m_ot6, "valore_recuperato"] = df.loc[m_ot6, "costo_perdita"]
df.loc[m_ot6, "perc_perdita"] = 0.0
df.loc[m_ot6, "nota_recovery"] = "Scarico interno (order_type=6)"

# Giacenza e Recesso: verifica se esiste una NDC (ordini_cliente.id_old = id_ordine)
# Se NON c'è NDC → valore_recuperato = costo_perdita (perdita 0)
# Se c'è NDC → lascia così (perdita reale)
m_giac_rec = df["tipo_rma"].isin(["Giacenza", "Recesso", "Reclamo e contestazioni"])
if m_giac_rec.any():
    gr_orders = df.loc[m_giac_rec, "id_ordine"].dropna().astype(int).unique().tolist()
    ndc_set = set()
    if gr_orders:
        gr_chunks = [gr_orders[i:i+500] for i in range(0, len(gr_orders), 500)]
        for chunk in gr_chunks:
            ids_str = ",".join(str(x) for x in chunk)
            q_ndc = f"SELECT DISTINCT id_old FROM ordini_cliente WHERE id_old IN ({ids_str})"
            df_ndc = pd.read_sql(text(q_ndc), engine)
            ndc_set.update(df_ndc["id_old"].tolist())
    if "ha_ndc" not in df.columns:
        df["ha_ndc"] = False
    df.loc[m_giac_rec, "ha_ndc"] = df.loc[m_giac_rec, "id_ordine"].isin(ndc_set)

    # Senza NDC → azzera perdita
    m_no_ndc = m_giac_rec & ~df["ha_ndc"]
    df.loc[m_no_ndc, "valore_recuperato"] = df.loc[m_no_ndc, "costo_perdita"]
    df.loc[m_no_ndc, "perc_perdita"] = 0.0
    m_no_ndc_giac = m_no_ndc & (df["tipo_rma"] == "Giacenza") & (df["nota_recovery"] == "")
    df.loc[m_no_ndc_giac, "nota_recovery"] = "Giacenza senza NDC (recupero al costo)"
    m_no_ndc_rec = m_no_ndc & (df["tipo_rma"] == "Recesso") & (df["nota_recovery"] == "")
    df.loc[m_no_ndc_rec, "nota_recovery"] = "Recesso senza NDC (recupero al costo)"
    m_no_ndc_recl = m_no_ndc & (df["tipo_rma"] == "Reclamo e contestazioni") & (df["nota_recovery"] == "")
    df.loc[m_no_ndc_recl, "nota_recovery"] = "Reclamo senza NDC (recupero al costo)"

    # Con NDC → lascia, aggiorna nota
    m_ndc = m_giac_rec & df["ha_ndc"]
    df.loc[m_ndc, "nota_recovery"] = df.loc[m_ndc, "nota_recovery"].astype(str) + " [NDC presente]"

# Perdita netta: include il costo di spedizione di rientro come voce
# aggiuntiva (non incluso in costo_perdita, mostrato separatamente in colonna).
df["perdita_netta"] = (
    df["costo_perdita"]
    + df["costo_sped_rientro"].fillna(0)
    - df["valore_claims"]
    - df["valore_recuperato"]
).round(2)

# ── 5b. Salva su DB (UPSERT su yeppon_stats.resi_impatto_economico) ─────
# Storico cumulativo: niente TRUNCATE. Le righe nuove vengono inserite,
# quelle esistenti (stesso id_rma+id_p) vengono aggiornate. Le righe più
# vecchie della finestra PERIODO restano intatte → storico permanente.
# Richiede UNIQUE KEY (id_rma, id_p) sulla tabella.
print("5b/5 Salvataggio su DB (UPSERT)...", end="", flush=True)

db_cols = {
    "id_rma": "id_rma", "id_ordine": "id_ordine", "id_p": "id_p",
    "tipo_rma": "tipo_rma", "marketplace": "marketplace",
    "qty_rma": "quantita_rma", "costo_unitario": "costo_unitario",
    "prezzo_vendita_unit": "prezzo_vendita_unit",
    "valore_rimborso": "valore_rimborso", "perc_rimborso": "perc_rimborso",
    "costo_perdita": "costo_perdita", "valore_claims": "valore_claims",
    "valore_recuperato": "valore_recuperato", "perdita_netta": "perdita_netta",
    "costo_sped_rientro": "costo_sped_rientro",
    "ha_ingresso": "ha_ingresso", "nota_recovery": "nota_recovery",
    "data_rma": "data_rma",
}
df_db = df[list(db_cols.keys())].copy()
df_db = df_db.rename(columns=db_cols)
df_db["ha_ingresso"] = df_db["ha_ingresso"].astype(int)
df_db["ha_ndc"] = 0
if "ha_ndc" in df.columns:
    df_db["ha_ndc"] = df["ha_ndc"].astype(int).values

# Fill NaN for DB
for c in ["costo_unitario", "prezzo_vendita_unit", "valore_rimborso",
          "perc_rimborso", "costo_perdita", "valore_claims",
          "valore_recuperato", "perdita_netta", "costo_sped_rientro"]:
    df_db[c] = df_db[c].fillna(0)
df_db["nota_recovery"] = df_db["nota_recovery"].fillna("").astype(str)
df_db["marketplace"] = df_db["marketplace"].fillna("").astype(str)
df_db["data_rma"] = pd.to_datetime(df_db["data_rma"], errors="coerce")

# Costruzione UPSERT (INSERT ... ON DUPLICATE KEY UPDATE)
all_cols = list(df_db.columns)
update_cols = [c for c in all_cols if c not in ("id_rma", "id_p")]
col_names = ", ".join(f"`{c}`" for c in all_cols)
placeholders = ", ".join(["%s"] * len(all_cols))
set_clause = ", ".join(f"`{c}`=VALUES(`{c}`)" for c in update_cols)
upsert_sql = (
    f"INSERT INTO yeppon_stats.resi_impatto_economico ({col_names}) "
    f"VALUES ({placeholders}) "
    f"ON DUPLICATE KEY UPDATE {set_clause}"
)

# Conversione robusta delle righe (NaN → None, Timestamp → datetime)
rows = []
for _, r in df_db.iterrows():
    vals = []
    for c in all_cols:
        v = r[c]
        if pd.isna(v):
            vals.append(None)
        elif isinstance(v, pd.Timestamp):
            vals.append(v.to_pydatetime())
        else:
            vals.append(v)
    rows.append(tuple(vals))

BATCH = 500
n_inserted_or_updated = 0
with engine.connect() as conn:
    raw_conn = conn.connection
    cursor = raw_conn.cursor()
    for i in range(0, len(rows), BATCH):
        cursor.executemany(upsert_sql, rows[i:i + BATCH])
        n_inserted_or_updated += cursor.rowcount or 0
    raw_conn.commit()
    cursor.close()

# Nota: con ON DUPLICATE KEY UPDATE, MySQL conta 1 per insert e 2 per update,
# quindi rowcount non equivale al numero di righe processate.
print(f" {len(df_db)} righe processate (rowcount MySQL: {n_inserted_or_updated}) ({time.time()-t0:.1f}s)")

# ── 6. Excel output ─────────────────────────────────────────────────────
print("6/6 Generazione Excel...", end="", flush=True)

COLS = [
    ("id_p",                 "ID Prodotto",      12),
    ("codice",               "Codice",           16),
    ("nome_prodotto",        "Nome",             40),
    ("marca",                "Marca",            16),
    ("id_ordine",            "ID Ordine Orig.",  14),
    ("id_rma",               "ID RMA",           10),
    ("tipo_rma",             "Tipo RMA",         22),
    ("data_rma",             "Data RMA",         12),
    ("qty_rma",              "Qty",               6),
    ("marketplace",          "Marketplace",      14),
    ("prezzo_vendita_unit",  "Prezzo Vendita €", 14),
    ("costo_unitario",       "Costo Acq. €",    12),
    ("valore_rimborso",      "Rimborso Cliente €",16),
    ("perc_rimborso",        "% Rimborso",       10),
    ("costo_perdita",        "Costo Perdita €",  14),
    ("costo_sped_rientro",   "Costo Sped. Rientro €", 16),
    ("valore_claims",        "Claims Corriere €", 16),
    ("ha_ingresso",          "Ingresso?",        10),
    ("data_ingresso",        "Data Ingresso",    12),
    ("aeh_tipo",             "Ultimo Mov.",      10),
    ("aeh_id_ordine",        "Ordine Rivend.",   14),
    ("macrotipo_alloc",      "Macrotipo Alloc.", 14),
    ("barcode_alloc",        "Barcode Alloc.",   14),
    ("desc_alloc",           "Allocazione",      14),
    ("perc_perdita",         "% Perdita",        10),
    ("valore_recuperato",    "Valore Recuperato €",16),
    ("nota_recovery",        "Nota Recovery",    28),
    ("perdita_netta",        "Perdita Netta €",  14),
]

fields  = [c[0] for c in COLS]
headers = [c[1] for c in COLS]
widths  = [c[2] for c in COLS]

for f in fields:
    if f not in df.columns:
        df[f] = ""

EUR_FMT = '#,##0.00 €'
eur_fields = ["prezzo_vendita_unit", "costo_unitario", "valore_rimborso",
              "costo_perdita", "costo_sped_rientro", "valore_claims",
              "valore_recuperato", "perdita_netta"]
pct_fields = ["perc_rimborso", "perc_perdita"]

out_path = os.path.join(
    os.path.expanduser("~"), "Downloads",
    f"dettaglio_resi_completo_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
)

tipi_rma = sorted(df["tipo_rma"].dropna().unique())

def _format_sheet(ws, n_rows):
    """Apply formatting to a sheet."""
    F_HDR  = PatternFill("solid", fgColor="1F3864")
    T_HDR  = Font(bold=True, color="FFFFFF", size=10)
    F_SI   = PatternFill("solid", fgColor="C6EFCE")
    F_NO   = PatternFill("solid", fgColor="FFC7CE")
    F_GAIN = PatternFill("solid", fgColor="DDEBF7")
    F_LOSS = PatternFill("solid", fgColor="FCE4D6")

    ws.row_dimensions[1].height = 34
    for ci in range(1, len(COLS) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = F_HDR
        c.font = T_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]

    ing_col   = fields.index("ha_ingresso") + 1
    perd_col  = fields.index("perdita_netta") + 1
    recup_col = fields.index("valore_recuperato") + 1

    for ri in range(2, n_rows + 2):
        for col_name in eur_fields:
            ci = fields.index(col_name) + 1
            ws.cell(row=ri, column=ci).number_format = EUR_FMT
        for col_name in pct_fields:
            ci = fields.index(col_name) + 1
            ws.cell(row=ri, column=ci).number_format = '0%'

        cell_ing = ws.cell(row=ri, column=ing_col)
        if cell_ing.value == "Sì":
            cell_ing.fill = F_SI; cell_ing.font = Font(color="006100")
        else:
            cell_ing.fill = F_NO; cell_ing.font = Font(color="9C0006")

        v = ws.cell(row=ri, column=perd_col).value
        if v is not None:
            try:
                if float(v) > 0:
                    ws.cell(row=ri, column=perd_col).fill = F_LOSS
                    ws.cell(row=ri, column=perd_col).font = Font(color="C00000", bold=True)
                elif float(v) < 0:
                    ws.cell(row=ri, column=perd_col).fill = F_GAIN
                    ws.cell(row=ri, column=perd_col).font = Font(color="006100")
            except (ValueError, TypeError):
                pass

        vr = ws.cell(row=ri, column=recup_col).value
        if vr:
            try:
                if float(vr) > 0:
                    ws.cell(row=ri, column=recup_col).fill = F_GAIN
            except (ValueError, TypeError):
                pass

    # TOTALE row
    last_row = n_rows + 2
    ws.cell(row=last_row, column=1).value = "TOTALE"
    ws.cell(row=last_row, column=1).font = Font(bold=True, size=11)
    for col_name in eur_fields:
        ci = fields.index(col_name) + 1
        col_letter = get_column_letter(ci)
        ws.cell(row=last_row, column=ci).value = (
            f"=SUM({col_letter}2:{col_letter}{last_row-1})"
        )
        ws.cell(row=last_row, column=ci).number_format = EUR_FMT
        ws.cell(row=last_row, column=ci).font = Font(bold=True, size=11)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{n_rows+1}"

# Build summary data
summary_rows = []

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # One sheet per tipo RMA
    for tipo in tipi_rma:
        df_tipo = df[df["tipo_rma"] == tipo].copy()
        if df_tipo.empty:
            continue

        # Sheet name max 31 chars
        sheet_name = tipo[:31]

        out_tipo = df_tipo[fields].copy()
        out_tipo["ha_ingresso"] = out_tipo["ha_ingresso"].map({True: "Sì", False: "No"})
        out_tipo.columns = headers

        out_tipo.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _format_sheet(ws, len(out_tipo))

        # Stats for summary
        n = len(df_tipo)
        n_ing = int(df_tipo["ha_ingresso"].sum())
        n_aeh = int(df_tipo["aeh_tipo"].notna().sum())
        n_sold = int(((df_tipo["aeh_tipo"] == "S") & (
            df_tipo["order_type_rivendita"].isin([1, 5])
            | ((df_tipo["order_type_rivendita"] == 3) & (df_tipo.get("pixmania_rivendita", pd.Series(dtype=float)) == 52))
        )).sum())
        n_sede = int(((df_tipo["aeh_tipo"] == "C") & (df_tipo["macrotipo_alloc"].str.upper() == "SEDE")).sum())
        n_assist = int(((df_tipo["aeh_tipo"] == "C") & (df_tipo["macrotipo_alloc"].str.upper() == "ASSISTENZA")).sum())
        summary_rows.append({
            "Tipo RMA": tipo,
            "Righe": n,
            "Con Ingresso": n_ing,
            "Con AEH": n_aeh,
            "Rivenduti (S+1)": n_sold,
            "In Stock (SEDE)": n_sede,
            "Assistenza": n_assist,
            "Tot. Rimborso €": round(df_tipo["valore_rimborso"].sum(), 2),
            "Tot. Costo Perdita €": round(df_tipo["costo_perdita"].sum(), 2),
            "Tot. Sped. Rientro €": round(df_tipo["costo_sped_rientro"].sum(), 2),
            "Tot. Claims €": round(df_tipo["valore_claims"].sum(), 2),
            "Tot. Recuperato €": round(df_tipo["valore_recuperato"].sum(), 2),
            "Tot. Perdita Netta €": round(df_tipo["perdita_netta"].sum(), 2),
        })

        print(f"  {sheet_name}: {n} righe", flush=True)

    # ── Riepilogo sheet ──────────────────────────────────────────────────
    df_summary = pd.DataFrame(summary_rows)
    # Add totale row
    tot = df_summary.select_dtypes(include="number").sum()
    tot["Tipo RMA"] = "TOTALE"
    df_summary = pd.concat([df_summary, pd.DataFrame([tot])], ignore_index=True)

    df_summary.to_excel(writer, sheet_name="Riepilogo", index=False)
    ws_r = writer.sheets["Riepilogo"]

    F_HDR = PatternFill("solid", fgColor="1F3864")
    T_HDR = Font(bold=True, color="FFFFFF", size=10)

    ws_r.row_dimensions[1].height = 30
    sum_widths = [22, 8, 12, 10, 14, 14, 12, 16, 18, 18, 14, 16, 18]
    for ci in range(1, len(df_summary.columns) + 1):
        c = ws_r.cell(row=1, column=ci)
        c.fill = F_HDR
        c.font = T_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_r.column_dimensions[get_column_letter(ci)].width = sum_widths[ci-1] if ci <= len(sum_widths) else 14

    eur_sum_cols = [8, 9, 10, 11, 12]  # 1-based column indices for € values
    for ri in range(2, len(df_summary) + 2):
        for ci in eur_sum_cols:
            ws_r.cell(row=ri, column=ci).number_format = EUR_FMT

    # Bold TOTALE row
    last_r = len(df_summary) + 1
    for ci in range(1, len(df_summary.columns) + 1):
        ws_r.cell(row=last_r, column=ci).font = Font(bold=True, size=11)

    # Color perdita netta column
    pn_col = 12
    F_LOSS = PatternFill("solid", fgColor="FCE4D6")
    for ri in range(2, len(df_summary) + 1):
        v = ws_r.cell(row=ri, column=pn_col).value
        if v and v > 0:
            ws_r.cell(row=ri, column=pn_col).fill = F_LOSS
            ws_r.cell(row=ri, column=pn_col).font = Font(color="C00000", bold=True)

    ws_r.freeze_panes = "A2"

print(f"\nOK ({time.time()-t0:.1f}s)")

# Console summary
print(f"\n{'Tipo RMA':<28} {'Righe':>6} {'Costo Perd.':>14} {'Claims':>12} {'Recuperato':>12} {'Perdita':>14}")
print("-" * 90)
for row in summary_rows:
    print(f"  {row['Tipo RMA']:<26} {row['Righe']:>6} {row['Tot. Costo Perdita €']:>12,.2f} € {row['Tot. Claims €']:>10,.2f} € {row['Tot. Recuperato €']:>10,.2f} € {row['Tot. Perdita Netta €']:>12,.2f} €")
print("-" * 90)
print(f"  {'TOTALE':<26} {sum(r['Righe'] for r in summary_rows):>6} "
      f"{sum(r['Tot. Costo Perdita €'] for r in summary_rows):>12,.2f} € "
      f"{sum(r['Tot. Claims €'] for r in summary_rows):>10,.2f} € "
      f"{sum(r['Tot. Recuperato €'] for r in summary_rows):>10,.2f} € "
      f"{sum(r['Tot. Perdita Netta €'] for r in summary_rows):>12,.2f} €")

print(f"\nFile: {out_path}")
