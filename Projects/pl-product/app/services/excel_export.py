"""Export Excel del P&L Prodotti.

Replica la formattazione di export_pl_prodotti.py (header blu, colori condizionali
sul margine/delta sped/resi, freeze panes, autofilter).
"""
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Palette colori ──────────────────────────────────────────────────────────
F_HEADER = PatternFill("solid", fgColor="1F3864")
F_GREEN = PatternFill("solid", fgColor="C6EFCE")
F_YELLOW = PatternFill("solid", fgColor="FFEB9C")
F_RED = PatternFill("solid", fgColor="FFC7CE")
F_GREY = PatternFill("solid", fgColor="F2F2F2")
F_WHITE = PatternFill("solid", fgColor="FFFFFF")

T_HEADER = Font(bold=True, color="FFFFFF", size=11)
T_GREEN = Font(color="276221", size=11)
T_YELLOW = Font(color="9C5700", size=11)
T_RED = Font(color="9C0006", size=11)
T_NORMAL = Font(color="333333", size=11)
T_NOTE = Font(italic=True, color="888888", size=10)

A_CENTER = Alignment(horizontal="center", vertical="center")
A_RIGHT = Alignment(horizontal="right", vertical="center")
A_LEFT = Alignment(horizontal="left", vertical="center")

# ── Colonne (field, header, width, fmt, color_type) ─────────────────────────
COLS = [
    ("id_p", "ID Prodotto", 10, "txt", None),
    ("codice", "Codice", 14, "str", None),
    ("nome", "Nome Prodotto", 50, "str", None),
    ("marca", "Marca", 22, "str", None),
    ("disp_fornitore", "Disp. Fornitore", 16, "int", None),
    ("status_prodotto", "Status", 10, "status", None),
    ("bloccato", "Bloccato", 10, "yn", None),
    ("bloccato_fino", "Bloccato Fino", 14, "date", None),
    ("num_ordini", "N° Ordini", 11, "int", None),
    ("tot_pezzi", "Pezzi Venduti", 13, "int", None),
    ("tot_fatturato", "Fatturato (€)", 15, "eur", None),
    ("tot_margine", "Margine (€)", 15, "eur", "margine"),
    ("perc_margine", "% Margine", 12, "perc", "margine"),
    ("tot_sped_fatt", "Incasso Sped. (€)", 16, "eur", None),
    ("tot_sped_costi", "Costo Sped. (€)", 15, "eur", None),
    ("delta_sped", "Delta Sped. (€)", 15, "eur", "delta"),
    ("perc_delta_sped", "% Delta Sped.", 13, "perc", "delta"),
    ("qty_resi", "Qty Resi", 10, "int", None),
    ("perc_resi", "% Resi", 9, "perc", "resi"),
    ("imp_eco_resi", "Impatto Eco. Resi (€)", 22, "eur", "imp_resi"),
    ("imp_Danni", "Danni (€)", 15, "eur", None),
    ("imp_Danno_rientrato", "Danno rientrato (€)", 18, "eur", None),
    ("imp_Difettosi", "Difettosi (€)", 15, "eur", None),
    ("imp_Giacenza", "Giacenza (€)", 15, "eur", None),
    ("imp_Mancato_Ritiro", "Mancato Ritiro (€)", 17, "eur", None),
    ("imp_Prodotto_non_conforme", "Prodotto non conforme (€)", 22, "eur", None),
    ("imp_Recesso", "Recesso (€)", 15, "eur", None),
    ("imp_Reclamo_contestazioni", "Reclamo e contestazioni (€)", 24, "eur", None),
    ("imp_Smarrimenti", "Smarrimenti (€)", 16, "eur", None),
    ("margine_effettivo", "Margine Effettivo (€)", 20, "eur", "margine_eff"),
    ("perc_margine_eff", "% Margine Effettivo", 20, "perc", "margine_eff"),
]


def _cell_color(color_type: str, value: float):
    """Restituisce (fill, font) in base al tipo di metrica e al valore."""
    if color_type in ("margine", "margine_eff"):
        if value < 0:
            return F_RED, T_RED
        if value <= 10:
            return F_YELLOW, T_YELLOW
        return F_GREEN, T_GREEN
    if color_type == "delta":
        return (F_GREEN, T_GREEN) if value >= 0 else (F_RED, T_RED)
    if color_type == "resi":
        if value <= 2:
            return F_GREEN, T_GREEN
        if value <= 5:
            return F_YELLOW, T_YELLOW
        return F_RED, T_RED
    if color_type == "imp_resi":
        return (F_RED, T_RED) if value > 0 else (F_GREEN, T_GREEN)
    return None, None


def _format_status(x):
    try:
        return "Attivo" if int(x) == 1 else "Disattivo"
    except (ValueError, TypeError):
        return str(x) if x else ""


def _format_bloccato(x):
    try:
        return "Sì" if int(x) == 1 else "No"
    except (ValueError, TypeError):
        return str(x) if x else ""


def _write_sheet(ws, df: pd.DataFrame, periodo_giorni: int):
    # Header
    ws.row_dimensions[1].height = 44
    for cell in ws[1]:
        cell.fill = F_HEADER
        cell.font = T_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Larghezze colonne
    for i, (_, _, w, _, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    # Dati riga per riga
    nrows = len(df)
    for ri in range(2, nrows + 2):
        row_data = df.iloc[ri - 2]
        base_fill = F_GREY if ri % 2 == 0 else F_WHITE

        pct_map = {
            "margine": float(row_data.get("perc_margine", 0) or 0),
            "margine_eff": float(row_data.get("perc_margine_eff", 0) or 0),
            "delta": float(row_data.get("delta_sped", 0) or 0),
            "resi": float(row_data.get("perc_resi", 0) or 0),
            "imp_resi": float(row_data.get("imp_eco_resi", 0) or 0),
        }

        for ci, (_field, _, _, fmt, ctype) in enumerate(COLS, 1):
            cell = ws.cell(row=ri, column=ci)

            if fmt == "eur":
                cell.number_format = '€#,##0.00'
                cell.alignment = A_RIGHT
            elif fmt == "perc":
                cell.number_format = '0.00"%"'
                cell.alignment = A_RIGHT
            elif fmt == "int":
                cell.number_format = "#,##0"
                cell.alignment = A_CENTER
            elif fmt == "txt":
                cell.number_format = '@'
                cell.value = str(cell.value) if cell.value is not None else ""
                cell.alignment = A_LEFT
            elif fmt in ("status", "yn"):
                cell.alignment = A_CENTER
            elif fmt == "date":
                if cell.value and str(cell.value) not in ("", "NaT", "None", "nan"):
                    cell.number_format = "DD/MM/YYYY"
                else:
                    cell.value = ""
                cell.alignment = A_CENTER
            else:
                cell.alignment = A_LEFT

            if ctype:
                fill, font = _cell_color(ctype, pct_map[ctype])
                cell.fill = fill
                cell.font = font
            else:
                cell.fill = base_fill
                cell.font = T_NORMAL

    # Nota a piè foglio
    note_row = nrows + 3
    note = ws.cell(
        row=note_row, column=1,
        value=(
            f"* Dati: ultimi {periodo_giorni} giorni | "
            "Margine Effettivo = Margine + Delta Spedizioni − Impatto Economico Resi. "
            "Logica identica a prodotto-pl-page.asp."
        ),
    )
    note.font = T_NOTE
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row, end_column=len(COLS),
    )


def df_to_excel_bytes(df: pd.DataFrame, periodo_giorni: int = 180) -> bytes:
    """Serializza il DataFrame nel formato Excel finale e ritorna i bytes."""
    df = df.copy()

    # Trasformazioni descrittive su colonne testuali
    if "status_prodotto" in df.columns:
        df["status_prodotto"] = df["status_prodotto"].apply(_format_status)
    if "bloccato" in df.columns:
        df["bloccato"] = df["bloccato"].apply(_format_bloccato)

    # Garantisce tutte le colonne attese
    for f, _, _, _, _ in COLS:
        if f not in df.columns:
            df[f] = 0

    fields = [c[0] for c in COLS]
    headers = [c[1] for c in COLS]
    out = df[fields].copy()
    out.columns = headers

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="P&L Prodotti", index=False)
        ws = writer.sheets["P&L Prodotti"]
        _write_sheet(ws, df, periodo_giorni)

    buf.seek(0)
    return buf.read()
